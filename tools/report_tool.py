# MIT License
#
# Copyright (c) 2026 Aryan Chavan
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in all
# copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

"""
ArynoxTech AI Agent - Professional Report Generation Tool
=========================================================
Enterprise-grade report generation supporting PDF, Excel, CSV, charts,
HTML dashboards, templates, scheduling, and comparison reports.
"""

import io
import json
import logging
import re
import time
from collections import OrderedDict
from datetime import datetime
from pathlib import Path as PathType
from typing import Any, Dict, List, Optional

import pandas as pd

from tools.base_tool import BaseTool, ToolResult
from config.settings import BASE_DIR, DIRS

logger = logging.getLogger(__name__)


class ReportTool(BaseTool):
    """
    Enterprise report generation tool supporting multiple output formats.

    Actions:
        generate_pdf, generate_excel, generate_csv, generate_chart,
        generate_html_dashboard, generate_dashboard, schedule_report,
        list_reports, generate_from_template, compare_reports
    """

    name: str = "report_tool"
    description: str = "Generate professional reports in PDF, Excel, CSV, charts, HTML dashboard formats with templates, scheduling, and comparison"
    version: str = "2.0.0"

    def __init__(self) -> None:
        super().__init__()
        self._reports_dir = DIRS.get("assets", BASE_DIR / "assets") / "reports"
        self._reports_dir.mkdir(parents=True, exist_ok=True)
        self._templates_dir = self._reports_dir / "templates"
        self._templates_dir.mkdir(parents=True, exist_ok=True)
        self._schedules_file = self._reports_dir / "schedules.json"
        self._generated_reports: List[PathType] = []

        self.supported_actions = OrderedDict([
            ("generate_pdf", self._generate_pdf),
            ("generate_excel", self._generate_excel),
            ("generate_csv", self._generate_csv),
            ("generate_chart", self._generate_chart),
            ("generate_html_dashboard", self._generate_html_dashboard),
            ("generate_dashboard", self._generate_dashboard),
            ("schedule_report", self._schedule_report),
            ("list_reports", self._list_reports),
            ("generate_from_template", self._generate_from_template),
            ("compare_reports", self._compare_reports),
        ])

    async def execute(self, **kwargs: Any) -> ToolResult:
        start_time = time.time()
        action = kwargs.get("action", "generate_pdf")

        handler = self.supported_actions.get(action)
        if not handler:
            elapsed = (time.time() - start_time) * 1000
            return ToolResult.failure(
                f"Unknown action: {action}. Available: {', '.join(self.supported_actions)}",
                execution_time_ms=elapsed,
            )

        try:
            return await handler(kwargs, start_time)
        except Exception as e:
            logger.exception(f"Report tool error in action '{action}': {e}")
            elapsed = (time.time() - start_time) * 1000
            return ToolResult.error_result(
                f"Report generation failed: {e}",
                error=str(e),
                execution_time_ms=elapsed,
            )

    # ------------------------------------------------------------------ #
    #  Utility helpers                                                    #
    # ------------------------------------------------------------------ #

    @staticmethod
    def _safe_filename(title: str, ext: str) -> str:
        safe = re.sub(r'[^\w\s\-]', '_', title).strip().replace(' ', '_')[:60]
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{safe}_{ts}{ext}"

    @staticmethod
    def _get_dataframe(data: Any) -> pd.DataFrame:
        if isinstance(data, pd.DataFrame):
            return data
        if isinstance(data, dict):
            return pd.DataFrame([data])
        if isinstance(data, list) and data:
            return pd.DataFrame(data)
        return pd.DataFrame()

    @staticmethod
    def _resolve_palette(name: str) -> List[str]:
        palettes = {
            "default": ["#4472C4", "#ED7D31", "#A5A5A5", "#FFC000", "#5B9BD5", "#70AD47", "#264478", "#9B57A0"],
            "corporate": ["#1B3A5C", "#2E6DA4", "#4A8BC2", "#6BA5D1", "#8FC1E0", "#B0D4ED", "#D0E6F5", "#E8F1F8"],
            "pastel": ["#FFB3BA", "#BAFFC9", "#BAE1FF", "#FFFFBA", "#E8BAFF", "#FFD9BA", "#BAFFF5", "#FFC8E1"],
            "monochrome": ["#1a1a1a", "#333333", "#4d4d4d", "#666666", "#808080", "#999999", "#b3b3b3", "#cccccc"],
        }
        return palettes.get(name, palettes["default"])

    def _record_report(self, path: PathType) -> None:
        self._generated_reports.append(path)

    # ------------------------------------------------------------------ #
    #  1. generate_pdf                                                    #
    # ------------------------------------------------------------------ #

    async def _generate_pdf(self, kwargs: Dict, start_time: float) -> ToolResult:
        title = kwargs.get("title", "Report")
        author = kwargs.get("author", "ArynoxTech AI Agent")
        subject = kwargs.get("subject", "")
        keywords = kwargs.get("keywords", "")
        sections = kwargs.get("sections", [])
        data = kwargs.get("data", {})

        df = self._get_dataframe(data.get("table", []))

        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.colors import HexColor, white
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_JUSTIFY
            from reportlab.platypus import (
                SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
                Image, PageBreak,
            )

            filename = self._safe_filename(title, ".pdf")
            filepath = self._reports_dir / filename

            page_w, page_h = A4
            margin = 72

            def header_footer(canvas, doc):
                canvas.saveState()
                canvas.setStrokeColor(HexColor("#4472C4"))
                canvas.setLineWidth(1.5)
                canvas.line(margin, page_h - 40, page_w - margin, page_h - 40)
                canvas.setFont("Helvetica", 8)
                canvas.setFillColor(HexColor("#666666"))
                canvas.drawString(margin, page_h - 35, title[:80])
                canvas.drawRightString(page_w - margin, page_h - 35, datetime.now().strftime("%Y-%m-%d"))
                canvas.setStrokeColor(HexColor("#CCCCCC"))
                canvas.setLineWidth(0.5)
                canvas.line(margin, 40, page_w - margin, 40)
                canvas.setFont("Helvetica", 8)
                canvas.setFillColor(HexColor("#999999"))
                canvas.drawString(margin, 28, f"Generated by ArynoxTech AI Agent  |  {author}")
                canvas.drawRightString(page_w - margin, 28, f"Page {doc.page}")
                canvas.restoreState()

            doc = SimpleDocTemplate(
                str(filepath),
                pagesize=A4,
                leftMargin=margin,
                rightMargin=margin,
                topMargin=margin + 20,
                bottomMargin=margin + 10,
                title=title,
                author=author,
                subject=subject,
                keywords=keywords,
            )

            styles = getSampleStyleSheet()

            styles.add(ParagraphStyle(
                "ReportTitle", parent=styles["Title"],
                fontSize=26, leading=32, textColor=HexColor("#1B3A5C"),
                spaceAfter=6, alignment=TA_LEFT, fontName="Helvetica-Bold",
            ))
            styles.add(ParagraphStyle(
                "Subtitle", parent=styles["Normal"],
                fontSize=12, leading=16, textColor=HexColor("#666666"),
                spaceAfter=20, alignment=TA_LEFT,
            ))
            styles.add(ParagraphStyle(
                "SectionHeading", parent=styles["Heading1"],
                fontSize=16, leading=22, textColor=HexColor("#1B3A5C"),
                spaceBefore=18, spaceAfter=8, fontName="Helvetica-Bold",
            ))
            styles.add(ParagraphStyle(
                "SubHeading", parent=styles["Heading2"],
                fontSize=13, leading=18, textColor=HexColor("#4472C4"),
                spaceBefore=12, spaceAfter=6, fontName="Helvetica-Bold",
            ))
            styles.add(ParagraphStyle(
                "BodyText2", parent=styles["Normal"],
                fontSize=10, leading=14, textColor=HexColor("#333333"),
                spaceAfter=8, alignment=TA_JUSTIFY,
            ))
            styles.add(ParagraphStyle(
                "BulletText", parent=styles["Normal"],
                fontSize=10, leading=14, textColor=HexColor("#333333"),
                leftIndent=20, bulletIndent=8, spaceAfter=4,
            ))
            styles.add(ParagraphStyle(
                "CodeBlock", parent=styles["Normal"],
                fontSize=8, leading=11, textColor=HexColor("#1a1a1a"),
                backColor=HexColor("#F5F5F5"), leftIndent=12,
                borderPadding=8, spaceBefore=6, spaceAfter=6,
                fontName="Courier",
            ))
            styles.add(ParagraphStyle(
                "Caption", parent=styles["Normal"],
                fontSize=9, leading=12, textColor=HexColor("#888888"),
                spaceAfter=12, alignment=TA_CENTER,
            ))

            story = []

            # Title Page
            story.append(Spacer(1, 80))
            story.append(Paragraph(title, styles["ReportTitle"]))
            story.append(Spacer(1, 8))
            line_table = Table([[""]], colWidths=[400])
            line_table.setStyle(TableStyle([
                ("LINEBELOW", (0, 0), (-1, -1), 3, HexColor("#4472C4")),
            ]))
            story.append(line_table)
            story.append(Spacer(1, 12))
            story.append(Paragraph(f"Author: {author}", styles["Subtitle"]))
            if subject:
                story.append(Paragraph(f"Subject: {subject}", styles["Subtitle"]))
            story.append(Paragraph(f"Date: {datetime.now().strftime('%B %d, %Y  %H:%M')}", styles["Subtitle"]))
            story.append(Spacer(1, 16))
            if keywords:
                story.append(Paragraph(f"Keywords: {keywords}", styles["Caption"]))

            # Table of Contents
            if sections:
                story.append(PageBreak())
                story.append(Paragraph("Table of Contents", styles["SectionHeading"]))
                story.append(Spacer(1, 8))
                toc_data = []
                for i, sec in enumerate(sections, 1):
                    toc_data.append([
                        Paragraph(f"{i}. {sec.get('heading', '')}", styles["BodyText2"]),
                        Paragraph("", styles["BodyText2"]),
                    ])
                if toc_data:
                    toc_table = Table(toc_data, colWidths=[400, 80])
                    toc_table.setStyle(TableStyle([
                        ("LINEBELOW", (0, 0), (-1, -1), 0.5, HexColor("#E0E0E0")),
                        ("TOPPADDING", (0, 0), (-1, -1), 4),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ]))
                    story.append(toc_table)

            # Sections
            for sec in sections:
                story.append(PageBreak())
                heading = sec.get("heading", "")
                if heading:
                    story.append(Paragraph(heading, styles["SectionHeading"]))
                    story.append(Spacer(1, 4))

                content = sec.get("content", "")
                if content:
                    paragraphs = content.split("\n") if isinstance(content, str) else [str(content)]
                    in_code = False
                    for para in paragraphs:
                        para = para.strip()
                        if not para:
                            continue
                        if para.startswith("```"):
                            in_code = not in_code
                            continue
                        if in_code:
                            story.append(Paragraph(para, styles["CodeBlock"]))
                        elif para.startswith("## "):
                            story.append(Paragraph(para[3:], styles["SubHeading"]))
                        elif para.startswith("- ") or para.startswith("* "):
                            story.append(Paragraph(para[2:], styles["BulletText"], bulletText="\u2022"))
                        else:
                            story.append(Paragraph(para, styles["BodyText2"]))

                # Table
                table_data = sec.get("table", [])
                if table_data:
                    tdf = self._get_dataframe(table_data)
                    if not tdf.empty:
                        cols = list(tdf.columns)
                        tbl_rows = [cols]
                        for _, row in tdf.iterrows():
                            tbl_rows.append([str(v) if v is not None else "" for v in row])

                        col_widths = [min(480 // max(len(cols), 1), 120) for _ in cols] if cols else [400]
                        tbl = Table(tbl_rows, colWidths=col_widths, repeatRows=1)

                        style_cmds = [
                            ("BACKGROUND", (0, 0), (-1, 0), HexColor("#1B3A5C")),
                            ("TEXTCOLOR", (0, 0), (-1, 0), white),
                            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                            ("FONTSIZE", (0, 0), (-1, 0), 9),
                            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                            ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#CCCCCC")),
                            ("TOPPADDING", (0, 0), (-1, -1), 5),
                            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                            ("LEFTPADDING", (0, 0), (-1, -1), 6),
                            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                        ]
                        for i in range(1, len(tbl_rows)):
                            bg = HexColor("#F2F7FB") if i % 2 == 0 else white
                            style_cmds.append(("BACKGROUND", (0, i), (-1, i), bg))

                        tbl.setStyle(TableStyle(style_cmds))
                        story.append(Spacer(1, 8))
                        story.append(tbl)
                        story.append(Paragraph(f"Table: {heading} Data", styles["Caption"]))

                # Chart
                chart_data = sec.get("chart", None)
                if chart_data:
                    try:
                        import matplotlib
                        matplotlib.use("Agg")
                        import matplotlib.pyplot as plt

                        cdf = self._get_dataframe(chart_data)
                        if not cdf.empty:
                            fig, ax = plt.subplots(figsize=(6.5, 3.5))
                            numeric = cdf.select_dtypes(include=["number"]).columns.tolist()
                            if numeric:
                                cdf.iloc[:, 0] = cdf.iloc[:, 0].astype(str)
                                for col in numeric[:3]:
                                    ax.plot(cdf.iloc[:, 0], cdf[col], marker="o", label=col, linewidth=1.5)
                                ax.set_title(heading, fontsize=12, fontweight="bold")
                                ax.legend(fontsize=8)
                                ax.tick_params(axis="x", rotation=45, labelsize=8)
                                ax.grid(axis="y", alpha=0.3)
                                plt.tight_layout()

                                img_buf = io.BytesIO()
                                fig.savefig(img_buf, format="png", dpi=150, bbox_inches="tight")
                                plt.close(fig)
                                img_buf.seek(0)
                                img = Image(img_buf, width=400, height=220)
                                story.append(Spacer(1, 6))
                                story.append(img)
                                story.append(Paragraph(f"Chart: {heading}", styles["Caption"]))
                            else:
                                plt.close(fig)
                    except Exception:
                        logger.warning("Could not embed chart in PDF section", exc_info=True)

            # Data Appendix
            if not df.empty and not sections:
                story.append(PageBreak())
                story.append(Paragraph("Data Appendix", styles["SectionHeading"]))
                story.append(Spacer(1, 8))
                cols = list(df.columns)
                tbl_rows = [cols]
                for _, row in df.iterrows():
                    tbl_rows.append([str(v) if v is not None else "" for v in row])
                col_widths = [min(480 // max(len(cols), 1), 120) for _ in cols] if cols else [400]
                tbl = Table(tbl_rows, colWidths=col_widths, repeatRows=1)
                style_cmds = [
                    ("BACKGROUND", (0, 0), (-1, 0), HexColor("#1B3A5C")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 9),
                    ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                    ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#CCCCCC")),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
                for i in range(1, len(tbl_rows)):
                    bg = HexColor("#F2F7FB") if i % 2 == 0 else white
                    style_cmds.append(("BACKGROUND", (0, i), (-1, i), bg))
                tbl.setStyle(TableStyle(style_cmds))
                story.append(tbl)

            doc.build(story, onFirstPage=header_footer, onLaterPages=header_footer)
            self._record_report(filepath)

            elapsed = (time.time() - start_time) * 1000
            return ToolResult.success(
                f"PDF report generated: {filename}",
                data={
                    "path": str(filepath),
                    "filename": filename,
                    "size_kb": round(filepath.stat().st_size / 1024, 2),
                    "sections": len(sections),
                    "format": "pdf",
                },
                execution_time_ms=elapsed,
            )
        except ImportError as e:
            elapsed = (time.time() - start_time) * 1000
            return ToolResult.failure(f"PDF generation requires reportlab: {e}", execution_time_ms=elapsed)

    # ------------------------------------------------------------------ #
    #  2. generate_excel                                                  #
    # ------------------------------------------------------------------ #

    async def _generate_excel(self, kwargs: Dict, start_time: float) -> ToolResult:
        title = kwargs.get("title", "Excel_Report")
        sheets = kwargs.get("sheets", {})
        data = kwargs.get("data", [])
        sheet_name = kwargs.get("sheet_name", "Sheet1")
        chart_configs = kwargs.get("charts", [])

        try:
            from openpyxl import Workbook
            from openpyxl.styles import (
                Font, PatternFill, Alignment, Border, Side,
            )
            from openpyxl.utils import get_column_letter
            from openpyxl.chart import BarChart, LineChart, PieChart, Reference
            from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

            filename = self._safe_filename(title, ".xlsx")
            filepath = self._reports_dir / filename

            wb = Workbook()

            header_fill = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
            alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
            thin_border = Border(
                left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC"),
            )

            def style_sheet(ws, sdf):
                for col_idx, col_name in enumerate(sdf.columns, 1):
                    cell = ws.cell(row=1, column=col_idx, value=col_name)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = thin_border
                for row_idx in range(2, ws.max_row + 1):
                    for col_idx in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.border = thin_border
                        cell.alignment = Alignment(vertical="center")
                        if row_idx % 2 == 0:
                            cell.fill = alt_fill
                for col_idx in range(1, ws.max_column + 1):
                    max_len = 0
                    for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
                        val = str(row[0]) if row[0] is not None else ""
                        max_len = max(max_len, len(val))
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)
                ws.freeze_panes = "A2"

            if sheets and isinstance(sheets, dict):
                default_ws = wb.active
                wb.remove(default_ws)

                for sname, sdata in sheets.items():
                    sdf = self._get_dataframe(
                        sdata.get("data", sdata if isinstance(sdata, list) else sdata.get("table", []))
                    )
                    ws = wb.create_sheet(title=sname[:31])

                    if not sdf.empty:
                        for col_idx, col_name in enumerate(sdf.columns, 1):
                            ws.cell(row=1, column=col_idx, value=col_name)
                        for row_idx, (_, row) in enumerate(sdf.iterrows(), 2):
                            for col_idx, val in enumerate(row, 1):
                                ws.cell(row=row_idx, column=col_idx, value=val)

                        style_sheet(ws, sdf)

                        summary_row = ws.max_row + 2
                        ws.cell(row=summary_row, column=1, value="Summary").font = Font(bold=True, size=11)
                        for col_idx in range(1, ws.max_column + 1):
                            col_letter = get_column_letter(col_idx)
                            col_vals = []
                            for r in range(2, ws.max_row - 1):
                                cell_val = ws.cell(row=r, column=col_idx).value
                                if isinstance(cell_val, (int, float)):
                                    col_vals.append(cell_val)
                            if col_vals:
                                end_row = ws.max_row - 1
                                ws.cell(row=summary_row + 1, column=col_idx,
                                        value=f"=SUM({col_letter}2:{col_letter}{end_row})")
                                ws.cell(row=summary_row + 2, column=col_idx,
                                        value=f"=AVERAGE({col_letter}2:{col_letter}{end_row})")
                                ws.cell(row=summary_row + 3, column=col_idx,
                                        value=f"=MAX({col_letter}2:{col_letter}{end_row})")
                                ws.cell(row=summary_row + 4, column=col_idx,
                                        value=f"=MIN({col_letter}2:{col_letter}{end_row})")

                        ws.cell(row=summary_row + 1, column=1, value="Sum").font = Font(bold=True)
                        ws.cell(row=summary_row + 2, column=1, value="Average").font = Font(bold=True)
                        ws.cell(row=summary_row + 3, column=1, value="Max").font = Font(bold=True)
                        ws.cell(row=summary_row + 4, column=1, value="Min").font = Font(bold=True)

                        for col_idx in range(1, ws.max_column + 1):
                            col_letter = get_column_letter(col_idx)
                            col_range = f"{col_letter}2:{col_letter}{ws.max_row - 1}"
                            try:
                                ws.conditional_formatting.add(col_range, ColorScaleRule(
                                    start_type="min", start_color="F2F7FB",
                                    mid_type="percentile", mid_value=50, mid_color="5B9BD5",
                                    end_type="max", end_color="1B3A5C",
                                ))
                            except Exception:
                                pass
                            break

                        for col_idx in range(1, ws.max_column + 1):
                            col_letter = get_column_letter(col_idx)
                            col_range = f"{col_letter}2:{col_letter}{ws.max_row - 1}"
                            if col_idx == 2:
                                try:
                                    ws.conditional_formatting.add(col_range, DataBarRule(
                                        start_type="min", end_type="max",
                                        color="4472C4", showValue=True,
                                    ))
                                except Exception:
                                    pass
                                break

                        for chart_cfg in chart_configs:
                            ctype = chart_cfg.get("type", "bar")
                            chart_title = chart_cfg.get("title", f"Chart {sname}")
                            cat_col = chart_cfg.get("category_column", sdf.columns[0] if not sdf.empty else None)
                            val_col = chart_cfg.get("value_column", sdf.columns[1] if len(sdf.columns) > 1 else None)
                            if cat_col and val_col and cat_col in sdf.columns and val_col in sdf.columns:
                                val_idx = list(sdf.columns).index(val_col) + 1
                                cat_idx = list(sdf.columns).index(cat_col) + 1
                                data_ref = Reference(ws, min_col=val_idx, min_row=1, max_row=ws.max_row - 1)
                                cats_ref = Reference(ws, min_col=cat_idx, min_row=2, max_row=ws.max_row - 1)

                                if ctype == "bar":
                                    chart = BarChart()
                                elif ctype == "line":
                                    chart = LineChart()
                                elif ctype == "pie":
                                    chart = PieChart()
                                else:
                                    chart = BarChart()
                                chart.title = chart_title
                                chart.y_axis.title = val_col
                                chart.x_axis.title = cat_col
                                chart.style = 10
                                chart.add_data(data_ref, titles_from_data=True)
                                if cats_ref:
                                    chart.set_categories(cats_ref)
                                chart.width = 18
                                chart.height = 12
                                ws.add_chart(chart, f"F{summary_row}")
            else:
                df = self._get_dataframe(data)
                ws = wb.active
                ws.title = sheet_name[:31]

                if not df.empty:
                    for col_idx, col_name in enumerate(df.columns, 1):
                        ws.cell(row=1, column=col_idx, value=col_name)
                    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
                        for col_idx, val in enumerate(row, 1):
                            ws.cell(row=row_idx, column=col_idx, value=val)

                    style_sheet(ws, df)

            wb.save(str(filepath))
            self._record_report(filepath)

            elapsed = (time.time() - start_time) * 1000
            return ToolResult.success(
                f"Excel report generated: {filename}",
                data={
                    "path": str(filepath),
                    "filename": filename,
                    "size_kb": round(filepath.stat().st_size / 1024, 2),
                    "sheets": len(wb.sheetnames),
                    "format": "xlsx",
                },
                execution_time_ms=elapsed,
            )
        except ImportError as e:
            elapsed = (time.time() - start_time) * 1000
            return ToolResult.failure(f"Excel generation requires openpyxl: {e}", execution_time_ms=elapsed)

    # ------------------------------------------------------------------ #
    #  3. generate_csv                                                    #
    # ------------------------------------------------------------------ #

    async def _generate_csv(self, kwargs: Dict, start_time: float) -> ToolResult:
        title = kwargs.get("title", "data_export")
        data = kwargs.get("data", [])
        columns = kwargs.get("columns", None)
        delimiter = kwargs.get("delimiter", ",")
        quote_char = kwargs.get("quote_char", '"')
        encoding = kwargs.get("encoding", "utf-8-sig")
        include_header = kwargs.get("include_header", True)
        bom = kwargs.get("bom", True)

        df = self._get_dataframe(data)
        if df.empty:
            elapsed = (time.time() - start_time) * 1000
            return ToolResult.failure("No data provided for CSV export", execution_time_ms=elapsed)

        if columns:
            cols = [c for c in columns if c in df.columns]
            if cols:
                df = df[cols]

        try:
            filename = self._safe_filename(title, ".csv")
            filepath = self._reports_dir / filename

            csv_kwargs = {
                "index": False,
                "sep": delimiter,
                "quotechar": quote_char,
                "encoding": "utf-8-sig" if (bom and "utf-8" in encoding.lower()) else encoding,
                "header": include_header,
                "lineterminator": "\r\n",
            }

            df.to_csv(filepath, **csv_kwargs)
            self._record_report(filepath)

            elapsed = (time.time() - start_time) * 1000
            return ToolResult.success(
                f"CSV exported: {filename}",
                data={
                    "path": str(filepath),
                    "filename": filename,
                    "size_kb": round(filepath.stat().st_size / 1024, 2),
                    "rows": len(df),
                    "columns": list(df.columns),
                    "format": "csv",
                },
                execution_time_ms=elapsed,
            )
        except Exception as e:
            elapsed = (time.time() - start_time) * 1000
            return ToolResult.error_result(f"CSV export failed: {e}", error=str(e), execution_time_ms=elapsed)

    # ------------------------------------------------------------------ #
    #  4. generate_chart                                                  #
    # ------------------------------------------------------------------ #

    async def _generate_chart(self, kwargs: Dict, start_time: float) -> ToolResult:
        title = kwargs.get("title", "Chart")
        data = kwargs.get("data", [])
        chart_type = kwargs.get("chart_type", "bar")
        x_column = kwargs.get("x_column", None)
        y_columns = kwargs.get("y_columns", None)
        palette = kwargs.get("palette", "default")
        colors = self._resolve_palette(palette)
        width = kwargs.get("width", 10)
        height = kwargs.get("height", 6)
        dpi = kwargs.get("dpi", 300)
        fmt = kwargs.get("format", "png")
        show_annotation = kwargs.get("show_annotation", False)
        horizontal = kwargs.get("horizontal", False)

        df = self._get_dataframe(data)
        if df.empty:
            elapsed = (time.time() - start_time) * 1000
            return ToolResult.failure("No data provided for chart", execution_time_ms=elapsed)

        try:
            import matplotlib
            matplotlib.use("Agg")
            import matplotlib.pyplot as plt
            import numpy as np

            numeric_cols = df.select_dtypes(include=["number"]).columns.tolist()
            if x_column and x_column in df.columns:
                x_vals = df[x_column].astype(str).tolist()
            else:
                x_vals = df.index.tolist()

            if y_columns:
                y_cols = [c for c in y_columns if c in df.columns]
            else:
                y_cols = numeric_cols[:8]

            if not y_cols:
                y_cols = [df.columns[1] if len(df.columns) > 1 else df.columns[0]]

            n_series = len(y_cols)
            auto_width = max(8, min(n_series * 2.5, 16))
            auto_height = max(5, min(n_series * 1.5, 10))

            fig, ax = plt.subplots(figsize=(width or auto_width, height or auto_height))

            x_pos = np.arange(len(x_vals))
            bar_width = 0.8 / max(n_series, 1)

            if chart_type == "bar":
                for i, col in enumerate(y_cols):
                    vals = df[col].values
                    pos = x_pos + (i - n_series / 2) * bar_width + bar_width / 2
                    bars = ax.bar(
                        pos, vals, bar_width * 0.9,
                        label=col, color=colors[i % len(colors)],
                        edgecolor="white", linewidth=0.5,
                    )
                    if show_annotation:
                        for bar in bars:
                            h = bar.get_height()
                            ax.annotate(f"{h:.1f}", (bar.get_x() + bar.get_width() / 2, h),
                                        ha="center", va="bottom", fontsize=7)
                if horizontal:
                    ax.invert_xaxis()

            elif chart_type == "bar_stacked":
                bottom = np.zeros(len(x_vals))
                for i, col in enumerate(y_cols):
                    vals = df[col].values
                    ax.bar(x_pos, vals, bar_width * n_series, bottom=bottom,
                           label=col, color=colors[i % len(colors)], edgecolor="white")
                    bottom += vals

            elif chart_type == "line":
                for i, col in enumerate(y_cols):
                    ax.plot(x_vals, df[col].values, marker="o", label=col,
                            color=colors[i % len(colors)], linewidth=2, markersize=5)
                if show_annotation:
                    for i, col in enumerate(y_cols):
                        for xi, yi in zip(x_vals, df[col].values):
                            ax.annotate(f"{yi:.1f}", (xi, yi), textcoords="offset points",
                                        xytext=(0, 8), ha="center", fontsize=7)

            elif chart_type == "pie":
                vals = df[y_cols[0]].values
                labels = x_vals if x_vals else df.index.tolist()
                ax.pie(
                    vals, labels=labels, autopct="%1.1f%%",
                    colors=colors[:len(labels)], startangle=90,
                    wedgeprops={"edgecolor": "white", "linewidth": 1},
                )

            elif chart_type == "scatter":
                for i, col in enumerate(y_cols):
                    ax.scatter(range(len(df)), df[col].values, label=col, s=60,
                               color=colors[i % len(colors)], alpha=0.7, edgecolors="white")

            elif chart_type == "area":
                for i, col in enumerate(y_cols):
                    ax.fill_between(range(len(df)), df[col].values, alpha=0.4,
                                    color=colors[i % len(colors)], label=col)
                    ax.plot(range(len(df)), df[col].values, color=colors[i % len(colors)], linewidth=1.5)

            elif chart_type == "histogram":
                bins = kwargs.get("bins", 20)
                for i, col in enumerate(y_cols):
                    ax.hist(df[col].dropna(), bins=bins, alpha=0.6,
                            color=colors[i % len(colors)], label=col, edgecolor="white")

            elif chart_type == "box":
                box_data = [df[col].dropna().values for col in y_cols]
                bp = ax.boxplot(box_data, labels=y_cols, patch_artist=True)
                for patch, color in zip(bp["boxes"], colors[:len(y_cols)]):
                    patch.set_facecolor(color)
                    patch.set_alpha(0.7)

            elif chart_type == "heatmap":
                heatmap_data = df.select_dtypes(include=["number"])
                if heatmap_data.shape[1] >= 2:
                    corr = heatmap_data.corr()
                    im = ax.imshow(corr, cmap="Blues", aspect="auto", vmin=-1, vmax=1)
                    plt.colorbar(im, ax=ax)
                    ax.set_xticks(range(len(corr.columns)))
                    ax.set_yticks(range(len(corr.columns)))
                    ax.set_xticklabels(corr.columns, rotation=45, ha="right", fontsize=8)
                    ax.set_yticklabels(corr.columns, fontsize=8)
                    for i in range(len(corr.columns)):
                        for j in range(len(corr.columns)):
                            ax.text(j, i, f"{corr.iloc[i, j]:.2f}", ha="center", va="center", fontsize=7)

            if chart_type not in ("pie", "heatmap"):
                ax.set_xlabel(x_column or "Index", fontsize=11)
                ax.set_ylabel(y_cols[0] if len(y_cols) == 1 else "", fontsize=11)
                if len(y_cols) > 1:
                    ax.legend(fontsize=9, framealpha=0.9, edgecolor="#CCCCCC")
                ax.grid(axis="y" if chart_type != "histogram" else "both", alpha=0.3, linestyle="--")
                ax.spines["top"].set_visible(False)
                ax.spines["right"].set_visible(False)
                ax.tick_params(axis="x", rotation=45 if len(x_vals) > 5 else 0, labelsize=9)

            ax.set_title(title, fontsize=14, fontweight="bold", pad=15, color="#1B3A5C")
            plt.tight_layout()

            ext = f".{fmt}"
            filename = self._safe_filename(title, ext)
            filepath = self._reports_dir / filename
            fig.savefig(filepath, dpi=dpi, bbox_inches="tight", format=fmt, facecolor="white")
            plt.close(fig)
            self._record_report(filepath)

            elapsed = (time.time() - start_time) * 1000
            return ToolResult.success(
                f"Chart generated: {filename}",
                data={
                    "path": str(filepath),
                    "filename": filename,
                    "size_kb": round(filepath.stat().st_size / 1024, 2),
                    "chart_type": chart_type,
                    "series": len(y_cols),
                    "format": fmt,
                },
                execution_time_ms=elapsed,
            )
        except ImportError as e:
            elapsed = (time.time() - start_time) * 1000
            return ToolResult.failure(f"Chart generation requires matplotlib: {e}", execution_time_ms=elapsed)
        except Exception as e:
            plt.close("all")
            elapsed = (time.time() - start_time) * 1000
            return ToolResult.error_result(f"Chart generation failed: {e}", error=str(e), execution_time_ms=elapsed)

    # ------------------------------------------------------------------ #
    #  5. generate_html_dashboard                                         #
    # ------------------------------------------------------------------ #

    async def _generate_html_dashboard(self, kwargs: Dict, start_time: float) -> ToolResult:
        title = kwargs.get("title", "Dashboard")
        kpis = kwargs.get("kpis", [])
        charts = kwargs.get("charts", [])
        tables = kwargs.get("tables", [])
        pages = kwargs.get("pages", [])

        try:
            import matplotlib
            matplotlib.use("Agg")
            import matplotlib.pyplot as plt
            import base64

            def fig_to_b64(fig):
                buf = io.BytesIO()
                fig.savefig(buf, format="png", dpi=150, bbox_inches="tight", facecolor="white")
                buf.seek(0)
                b64 = base64.b64encode(buf.read()).decode("utf-8")
                plt.close(fig)
                return b64

            kpi_cards = ""
            for i, kpi in enumerate(kpis):
                label = kpi.get("label", f"KPI {i+1}")
                value = kpi.get("value", "")
                change = kpi.get("change", "")
                icon = kpi.get("icon", "&#128202;")
                arrow = "&#9650;" if change and not str(change).startswith("-") else "&#9660;"
                color = "#27AE60" if not str(change).startswith("-") else "#E74C3C"
                kpi_cards += f"""
                <div class="kpi-card">
                    <div class="kpi-icon">{icon}</div>
                    <div class="kpi-content">
                        <div class="kpi-label">{label}</div>
                        <div class="kpi-value">{value}</div>
                        <div class="kpi-change" style="color:{color}">{arrow} {change}</div>
                    </div>
                </div>"""

            charts_html = ""
            for i, chart_cfg in enumerate(charts):
                chart_data = chart_cfg.get("data", [])
                chart_type = chart_cfg.get("type", "bar")
                chart_title = chart_cfg.get("title", f"Chart {i+1}")
                cdf = self._get_dataframe(chart_data)
                chart_img = ""
                if not cdf.empty:
                    fig, ax = plt.subplots(figsize=(6, 3.5))
                    numeric = cdf.select_dtypes(include=["number"]).columns.tolist()
                    if numeric:
                        if chart_type in ("bar", "line"):
                            cats = cdf.iloc[:, 0].astype(str) if cdf.columns[0] not in numeric else cdf.index.astype(str)
                            for j, col in enumerate(numeric[:3]):
                                if chart_type == "bar":
                                    ax.bar(cats, cdf[col], alpha=0.8, label=col)
                                else:
                                    ax.plot(cats, cdf[col], marker="o", label=col)
                            ax.legend(fontsize=8)
                            ax.tick_params(axis="x", rotation=45, labelsize=8)
                        elif chart_type == "pie":
                            ax.pie(cdf[numeric[0]], labels=cdf.iloc[:, 0].astype(str),
                                   autopct="%1.1f%%", startangle=90)
                        ax.set_title(chart_title, fontsize=12, fontweight="bold")
                        ax.grid(axis="y", alpha=0.3)
                        plt.tight_layout()
                        chart_img = fig_to_b64(fig)
                    else:
                        plt.close(fig)

                if chart_img:
                    charts_html += f"""
                <div class="card chart-card">
                    <h3>{chart_title}</h3>
                    <img src="data:image/png;base64,{chart_img}" alt="{chart_title}" style="width:100%;max-width:700px;">
                </div>"""

            tables_html = ""
            for i, tbl_cfg in enumerate(tables):
                tbl_data = tbl_cfg.get("data", [])
                tbl_title = tbl_cfg.get("title", f"Table {i+1}")
                tdf = self._get_dataframe(tbl_data)
                if not tdf.empty:
                    tbl_headers = ""
                    for col in tdf.columns:
                        tbl_headers += f"<th>{col}</th>"
                    tbl_rows = ""
                    for _, row in tdf.iterrows():
                        cells = ""
                        for val in row:
                            cells += f"<td>{val}</td>"
                        tbl_rows += f"<tr>{cells}</tr>"
                    tables_html += f"""
                <div class="card table-card">
                    <h3>{tbl_title}</h3>
                    <div class="table-wrapper">
                        <table><thead><tr>{tbl_headers}</tr></thead><tbody>{tbl_rows}</tbody></table>
                    </div>
                </div>"""

            tabs_html = ""
            pages_content = ""
            if pages:
                tabs_html = '<div class="tab-nav">'
                for pi, page in enumerate(pages):
                    active = " active" if pi == 0 else ""
                    tabs_html += f"""<button class="tab-btn{active}" onclick="showTab({pi})">{page.get("title", f"Page {pi+1}")}</button>"""
                    page_charts_html = ""
                    for pc in page.get("charts", []):
                        pcdf = self._get_dataframe(pc.get("data", []))
                        pc_img = ""
                        if not pcdf.empty:
                            fig, ax = plt.subplots(figsize=(5, 3))
                            numeric = pcdf.select_dtypes(include=["number"]).columns.tolist()
                            if numeric:
                                ax.plot(pcdf.iloc[:, 0].astype(str), pcdf[numeric[0]], marker="o")
                                ax.set_title(pc.get("title", ""), fontsize=11)
                                ax.grid(alpha=0.3)
                                plt.tight_layout()
                                pc_img = fig_to_b64(fig)
                            else:
                                plt.close(fig)
                        if pc_img:
                            page_charts_html += f"""<img src="data:image/png;base64,{pc_img}" style="max-width:100%;margin:10px 0;">"""
                    hidden = " hidden" if pi != 0 else ""
                    pages_content += f"""
                    <div class="tab-content{hidden}" id="tab{pi}">
                        <h3>{page.get("title", "")}</h3>
                        <p>{page.get("content", "")}</p>
                        {page_charts_html}
                    </div>"""
                tabs_html += "</div>"

            timestamp = datetime.now().strftime("%B %d, %Y at %H:%M")
            html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>{title}</title>
<style>
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ font-family:'Segoe UI',system-ui,-apple-system,sans-serif; background:#f5f7fa; color:#333; }}
.dashboard {{ max-width:1400px; margin:0 auto; padding:20px; }}
.header {{ background:linear-gradient(135deg,#1B3A5C,#2E6DA4); color:white; padding:30px; border-radius:12px; margin-bottom:24px; }}
.header h1 {{ font-size:28px; font-weight:600; }}
.header p {{ opacity:0.85; margin-top:6px; font-size:14px; }}
.kpi-grid {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(220px,1fr)); gap:16px; margin-bottom:24px; }}
.kpi-card {{ background:white; border-radius:10px; padding:20px; display:flex; align-items:center; gap:16px; box-shadow:0 2px 8px rgba(0,0,0,0.06); transition:transform .15s; }}
.kpi-card:hover {{ transform:translateY(-2px); box-shadow:0 4px 16px rgba(0,0,0,0.1); }}
.kpi-icon {{ font-size:32px; width:50px; height:50px; display:flex; align-items:center; justify-content:center; background:#f0f4f8; border-radius:10px; }}
.kpi-content {{ flex:1; }}
.kpi-label {{ font-size:12px; color:#888; text-transform:uppercase; letter-spacing:.5px; }}
.kpi-value {{ font-size:24px; font-weight:700; margin:4px 0; }}
.kpi-change {{ font-size:13px; font-weight:600; }}
.card {{ background:white; border-radius:10px; padding:20px; margin-bottom:20px; box-shadow:0 2px 8px rgba(0,0,0,0.06); }}
.card h3 {{ font-size:16px; font-weight:600; color:#1B3A5C; margin-bottom:12px; }}
.chart-grid {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(450px,1fr)); gap:20px; }}
.table-wrapper {{ overflow-x:auto; }}
table {{ width:100%; border-collapse:collapse; font-size:13px; }}
th {{ background:#1B3A5C; color:white; padding:10px 12px; text-align:left; font-weight:600; }}
td {{ padding:8px 12px; border-bottom:1px solid #eee; }}
tr:nth-child(even) {{ background:#f8fafc; }}
tr:hover {{ background:#eef3f9; }}
.tab-nav {{ display:flex; gap:4px; margin-bottom:16px; background:white; border-radius:8px; padding:4px; box-shadow:0 1px 4px rgba(0,0,0,0.06); }}
.tab-btn {{ flex:1; padding:10px 16px; border:none; background:transparent; border-radius:6px; font-size:14px; font-weight:500; cursor:pointer; color:#666; transition:all .15s; }}
.tab-btn:hover {{ background:#f0f4f8; }}
.tab-btn.active {{ background:#1B3A5C; color:white; }}
.tab-content {{ animation:fade .2s; }}
.tab-content.hidden {{ display:none; }}
@keyframes fade {{ from{{opacity:0}} to{{opacity:1}} }}
.footer {{ text-align:center; color:#aaa; font-size:12px; padding:20px; }}
.export-bar {{ display:flex; gap:8px; margin-bottom:16px; flex-wrap:wrap; }}
@media(max-width:768px){{ .kpi-grid{{grid-template-columns:1fr 1fr}} .chart-grid{{grid-template-columns:1fr}} .header{{padding:20px}} }}
</style>
</head>
<body>
<div class="dashboard">
    <div class="header">
        <h1>{title}</h1>
        <p>Generated on {timestamp} | ArynoxTech AI Agent</p>
        <div class="export-bar" style="margin-top:12px;">
            <button class="btn" style="background:#1B3A5C;color:white;padding:8px 16px;border:none;border-radius:6px;cursor:pointer;" onclick="window.print()">Print / PDF</button>
        </div>
    </div>
    <div class="kpi-grid">{kpi_cards}</div>
    {tabs_html}
    {pages_content}
    <div class="chart-grid">{charts_html}</div>
    {tables_html}
    <div class="footer">Generated by ArynoxTech AI Agent Report Tool v{self.version}</div>
</div>
<script>
function showTab(idx){{ document.querySelectorAll('.tab-content').forEach(t=>t.classList.add('hidden')); document.querySelectorAll('.tab-btn').forEach(b=>b.classList.remove('active')); document.getElementById('tab'+idx).classList.remove('hidden'); document.querySelectorAll('.tab-btn')[idx].classList.add('active'); }}
</script>
</body>
</html>"""

            filename = self._safe_filename(title, ".html")
            filepath = self._reports_dir / filename
            filepath.write_text(html, encoding="utf-8")
            self._record_report(filepath)

            elapsed = (time.time() - start_time) * 1000
            return ToolResult.success(
                f"HTML dashboard generated: {filename}",
                data={
                    "path": str(filepath),
                    "filename": filename,
                    "size_kb": round(filepath.stat().st_size / 1024, 2),
                    "kpis": len(kpis),
                    "charts": len(charts),
                    "tables": len(tables),
                    "pages": len(pages),
                    "format": "html",
                },
                execution_time_ms=elapsed,
            )
        except ImportError as e:
            elapsed = (time.time() - start_time) * 1000
            return ToolResult.failure(f"Dashboard generation requires matplotlib: {e}", execution_time_ms=elapsed)
        except Exception as e:
            plt.close("all")
            elapsed = (time.time() - start_time) * 1000
            return ToolResult.error_result(f"Dashboard generation failed: {e}", error=str(e), execution_time_ms=elapsed)

    # ------------------------------------------------------------------ #
    #  6. generate_dashboard (multi-chart combined)                       #
    # ------------------------------------------------------------------ #

    async def _generate_dashboard(self, kwargs: Dict, start_time: float) -> ToolResult:
        title = kwargs.get("title", "Dashboard")
        charts = kwargs.get("charts", [])
        data = kwargs.get("data", {})

        try:
            import matplotlib
            matplotlib.use("Agg")
            import matplotlib.pyplot as plt
            import numpy as np

            df = self._get_dataframe(data)
            n_charts = len(charts)
            if n_charts == 0 and not df.empty:
                numeric = df.select_dtypes(include=["number"]).columns.tolist()
                cat_col = df.columns[0] if df.columns[0] not in numeric else None
                for col in numeric[:4]:
                    charts.append({"type": "bar", "title": col, "x_column": cat_col, "y_columns": [col]})
                n_charts = len(charts)

            if n_charts == 0:
                elapsed = (time.time() - start_time) * 1000
                return ToolResult.failure("No charts or data provided", execution_time_ms=elapsed)

            cols = 2 if n_charts <= 4 else 3
            rows = (n_charts + cols - 1) // cols

            fig, axes = plt.subplots(rows, cols, figsize=(cols * 5, rows * 3.5))
            axes_flat = axes.flatten() if isinstance(axes, np.ndarray) else [axes]

            palette_colors = self._resolve_palette(kwargs.get("palette", "default"))

            for idx, (ax, chart_cfg) in enumerate(zip(axes_flat, charts)):
                chart_type = chart_cfg.get("type", "bar")
                chart_title = chart_cfg.get("title", f"Chart {idx+1}")
                x_col = chart_cfg.get("x_column", None)
                y_cols = chart_cfg.get("y_columns", None)

                cdf = self._get_dataframe(chart_cfg.get("data", df if not df.empty else {}))

                if cdf.empty:
                    ax.text(0.5, 0.5, "No data", ha="center", va="center", transform=ax.transAxes)
                    continue

                numeric = cdf.select_dtypes(include=["number"]).columns.tolist()
                if x_col and x_col in cdf.columns:
                    x_vals = cdf[x_col].astype(str)
                else:
                    x_vals = cdf.index.astype(str)

                if y_cols:
                    y_list = [c for c in y_cols if c in cdf.columns]
                else:
                    y_list = numeric[:3]

                if not y_list:
                    y_list = [cdf.columns[-1]]

                if chart_type == "bar":
                    for j, col in enumerate(y_list):
                        ax.bar(x_vals, cdf[col], alpha=0.8, color=palette_colors[j % len(palette_colors)], label=col)
                    ax.legend(fontsize=8)
                    ax.tick_params(axis="x", rotation=45, labelsize=7)
                elif chart_type == "line":
                    for j, col in enumerate(y_list):
                        ax.plot(x_vals, cdf[col], marker="o", color=palette_colors[j % len(palette_colors)], label=col)
                    ax.legend(fontsize=8)
                elif chart_type == "pie":
                    ax.pie(cdf[y_list[0]], labels=x_vals.tolist(), autopct="%1.1f%%",
                           colors=palette_colors[:len(x_vals)], startangle=90)
                elif chart_type == "scatter":
                    for j, col in enumerate(y_list):
                        ax.scatter(range(len(cdf)), cdf[col], color=palette_colors[j % len(palette_colors)], label=col, s=30)
                elif chart_type == "area":
                    for j, col in enumerate(y_list):
                        ax.fill_between(range(len(cdf)), cdf[col], alpha=0.3, color=palette_colors[j % len(palette_colors)], label=col)
                elif chart_type == "histogram":
                    for j, col in enumerate(y_list):
                        ax.hist(cdf[col].dropna(), bins=15, alpha=0.6, color=palette_colors[j % len(palette_colors)], label=col)
                elif chart_type == "box":
                    box_data = [cdf[col].dropna().values for col in y_list]
                    bp = ax.boxplot(box_data, labels=y_list, patch_artist=True)
                    for patch, color in zip(bp["boxes"], palette_colors):
                        patch.set_facecolor(color)
                        patch.set_alpha(0.7)
                elif chart_type == "heatmap":
                    corr = cdf.select_dtypes(include=["number"]).corr()
                    ax.imshow(corr, cmap="Blues", aspect="auto")
                    ax.set_xticks(range(len(corr.columns)))
                    ax.set_yticks(range(len(corr.columns)))
                    ax.set_xticklabels(corr.columns, rotation=45, ha="right", fontsize=7)
                    ax.set_yticklabels(corr.columns, fontsize=7)

                ax.set_title(chart_title, fontsize=10, fontweight="bold")
                if chart_type not in ("pie", "heatmap"):
                    ax.grid(axis="y", alpha=0.2)
                    ax.spines["top"].set_visible(False)
                    ax.spines["right"].set_visible(False)

            for idx in range(n_charts, len(axes_flat)):
                axes_flat[idx].set_visible(False)

            fig.suptitle(title, fontsize=16, fontweight="bold", y=1.02, color="#1B3A5C")
            plt.tight_layout()

            filename = self._safe_filename(title, ".png")
            filepath = self._reports_dir / filename
            fig.savefig(filepath, dpi=300, bbox_inches="tight", facecolor="white")
            plt.close(fig)
            self._record_report(filepath)

            combined = [{"path": str(filepath), "filename": filename, "type": "png"}]

            try:
                from weasyprint import HTML
                pdf_filename = self._safe_filename(title, ".pdf")
                pdf_path = self._reports_dir / pdf_filename
                HTML(string=f"<html><body><img src='file:///{filepath}' style='max-width:100%'/></body></html>").write_pdf(str(pdf_path))
                self._record_report(pdf_path)
                combined.append({"path": str(pdf_path), "filename": pdf_filename, "type": "pdf"})
            except Exception:
                pass

            elapsed = (time.time() - start_time) * 1000
            return ToolResult.success(
                f"Dashboard generated: {filename}" + (f" + PDF" if len(combined) > 1 else ""),
                data={
                    "files": combined,
                    "charts": n_charts,
                    "layout": f"{rows}x{cols}",
                    "format": "png",
                },
                execution_time_ms=elapsed,
            )
        except ImportError as e:
            elapsed = (time.time() - start_time) * 1000
            return ToolResult.failure(f"Dashboard requires matplotlib: {e}", execution_time_ms=elapsed)
        except Exception as e:
            plt.close("all")
            elapsed = (time.time() - start_time) * 1000
            return ToolResult.error_result(f"Dashboard generation failed: {e}", error=str(e), execution_time_ms=elapsed)

    # ------------------------------------------------------------------ #
    #  7. schedule_report                                                #
    # ------------------------------------------------------------------ #

    async def _schedule_report(self, kwargs: Dict, start_time: float) -> ToolResult:
        action = kwargs.get("schedule_action", "create")
        schedule_id = kwargs.get("schedule_id", None)

        try:
            schedules = []
            if self._schedules_file.exists():
                schedules = json.loads(self._schedules_file.read_text(encoding="utf-8"))

            if action == "create":
                new_schedule = {
                    "id": schedule_id or f"sched_{int(time.time())}",
                    "cron": kwargs.get("cron", ""),
                    "interval": kwargs.get("interval", ""),
                    "report_config": {
                        "action": kwargs.get("report_action", "generate_pdf"),
                        "title": kwargs.get("title", "Scheduled Report"),
                        "format": kwargs.get("format", "pdf"),
                        "data_source": kwargs.get("data_source", ""),
                        "template": kwargs.get("template", ""),
                    },
                    "created_at": datetime.now().isoformat(),
                    "next_run": "",
                    "enabled": True,
                }
                schedules.append(new_schedule)
                self._schedules_file.write_text(
                    json.dumps(schedules, indent=2, default=str), encoding="utf-8"
                )

                elapsed = (time.time() - start_time) * 1000
                return ToolResult.success(
                    f"Report schedule created: {new_schedule['id']}",
                    data={"schedule": new_schedule, "schedules_count": len(schedules)},
                    execution_time_ms=elapsed,
                )

            elif action == "list":
                elapsed = (time.time() - start_time) * 1000
                return ToolResult.success(
                    f"{len(schedules)} schedule(s) found",
                    data={"schedules": schedules, "count": len(schedules)},
                    execution_time_ms=elapsed,
                )

            elif action == "cancel":
                sid = schedule_id or kwargs.get("schedule_id", "")
                schedules = [s for s in schedules if s.get("id") != sid]
                self._schedules_file.write_text(
                    json.dumps(schedules, indent=2, default=str), encoding="utf-8"
                )
                elapsed = (time.time() - start_time) * 1000
                return ToolResult.success(
                    f"Schedule cancelled: {sid}" if sid else "All schedules cleared",
                    data={"schedules_count": len(schedules)},
                    execution_time_ms=elapsed,
                )
            else:
                elapsed = (time.time() - start_time) * 1000
                return ToolResult.failure(f"Unknown schedule_action: {action}", execution_time_ms=elapsed)
        except Exception as e:
            elapsed = (time.time() - start_time) * 1000
            return ToolResult.error_result(f"Schedule operation failed: {e}", error=str(e), execution_time_ms=elapsed)

    # ------------------------------------------------------------------ #
    #  8. list_reports                                                   #
    # ------------------------------------------------------------------ #

    async def _list_reports(self, kwargs: Dict, start_time: float) -> ToolResult:
        name_pattern = kwargs.get("name_pattern", "")
        date_from = kwargs.get("date_from", "")
        date_to = kwargs.get("date_to", "")
        file_format = kwargs.get("format", "")

        allowed = (".pdf", ".xlsx", ".xls", ".csv", ".png", ".svg", ".jpg", ".html", ".json")

        try:
            reports = []
            for f in sorted(self._reports_dir.iterdir(), key=lambda p: p.stat().st_mtime, reverse=True):
                if not f.is_file() or f.suffix.lower() not in allowed:
                    continue
                if f.parent == self._templates_dir:
                    continue

                fname = f.name
                if name_pattern and name_pattern.lower() not in fname.lower():
                    continue
                if file_format and not fname.lower().endswith(f".{file_format.lower()}"):
                    continue

                mtime = datetime.fromtimestamp(f.stat().st_mtime)
                if date_from:
                    try:
                        if mtime < datetime.fromisoformat(date_from):
                            continue
                    except Exception:
                        pass
                if date_to:
                    try:
                        if mtime > datetime.fromisoformat(date_to):
                            continue
                    except Exception:
                        pass

                reports.append({
                    "filename": fname,
                    "size_kb": round(f.stat().st_size / 1024, 2),
                    "format": f.suffix.lower().lstrip("."),
                    "created": mtime.isoformat(),
                    "path": str(f),
                })

            elapsed = (time.time() - start_time) * 1000
            return ToolResult.success(
                f"{len(reports)} report(s) found",
                data={"reports": reports, "count": len(reports)},
                execution_time_ms=elapsed,
            )
        except Exception as e:
            elapsed = (time.time() - start_time) * 1000
            return ToolResult.error_result(f"Failed to list reports: {e}", error=str(e), execution_time_ms=elapsed)

    # ------------------------------------------------------------------ #
    #  9. generate_from_template                                          #
    # ------------------------------------------------------------------ #

    async def _generate_from_template(self, kwargs: Dict, start_time: float) -> ToolResult:
        template = kwargs.get("template", None)
        template_name = kwargs.get("template_name", "")
        variables = kwargs.get("variables", {})
        output_format = kwargs.get("format", "pdf")
        title = kwargs.get("title", "Template Report")

        try:
            import jinja2
            import yaml

            if template is None and template_name:
                tmpl_path = self._templates_dir / f"{template_name}.json"
                if not tmpl_path.exists():
                    tmpl_path = self._templates_dir / f"{template_name}.yaml"
                if not tmpl_path.exists():
                    tmpl_path = self._templates_dir / f"{template_name}.yml"
                if tmpl_path.exists():
                    raw = tmpl_path.read_text(encoding="utf-8")
                    if tmpl_path.suffix == ".json":
                        template = json.loads(raw)
                    else:
                        template = yaml.safe_load(raw)
                else:
                    elapsed = (time.time() - start_time) * 1000
                    return ToolResult.failure(f"Template not found: {template_name}", execution_time_ms=elapsed)

            if template is None:
                elapsed = (time.time() - start_time) * 1000
                return ToolResult.failure("No template provided", execution_time_ms=elapsed)

            if isinstance(template, str):
                template = json.loads(template)

            def render_value(val):
                if isinstance(val, str):
                    env = jinja2.Environment()
                    tpl = env.from_string(val)
                    return tpl.render(**variables)
                if isinstance(val, dict):
                    return {k: render_value(v) for k, v in val.items()}
                if isinstance(val, list):
                    return [render_value(v) for v in val]
                return val

            filled = render_value(template)

            sec_title = filled.get("title", title)
            sec_sections = []
            for sec in filled.get("sections", []):
                sec_type = sec.get("type", "text")
                section = {"heading": sec.get("heading", "")}
                if sec_type == "text":
                    section["content"] = sec.get("content", "")
                elif sec_type == "table":
                    section["table"] = sec.get("content", sec.get("data", []))
                elif sec_type == "chart":
                    section["chart"] = sec.get("content", sec.get("data", []))
                sec_sections.append(section)

            template_kwargs = {
                "title": sec_title,
                "author": variables.get("author", kwargs.get("author", "ArynoxTech AI Agent")),
                "sections": sec_sections,
                "data": kwargs.get("data", {}),
            }

            if output_format == "pdf":
                result = await self._generate_pdf(template_kwargs, start_time)
            elif output_format == "excel":
                result = await self._generate_excel(template_kwargs, start_time)
            elif output_format == "csv":
                result = await self._generate_csv(template_kwargs, start_time)
            elif output_format == "chart":
                result = await self._generate_chart(template_kwargs, start_time)
            elif output_format == "html":
                result = await self._generate_html_dashboard(template_kwargs, start_time)
            else:
                result = await self._generate_pdf(template_kwargs, start_time)

            return result

        except ImportError as e:
            elapsed = (time.time() - start_time) * 1000
            return ToolResult.failure(f"Template generation requires jinja2 and PyYAML: {e}", execution_time_ms=elapsed)
        except Exception as e:
            elapsed = (time.time() - start_time) * 1000
            return ToolResult.error_result(f"Template generation failed: {e}", error=str(e), execution_time_ms=elapsed)

    # ------------------------------------------------------------------ #
    # 10. compare_reports                                                #
    # ------------------------------------------------------------------ #

    async def _compare_reports(self, kwargs: Dict, start_time: float) -> ToolResult:
        data_a = kwargs.get("data_a", [])
        data_b = kwargs.get("data_b", [])
        key_column = kwargs.get("key_column", None)
        compare_columns = kwargs.get("compare_columns", None)
        title = kwargs.get("title", "Comparison Report")

        df_a = self._get_dataframe(data_a)
        df_b = self._get_dataframe(data_b)

        if df_a.empty or df_b.empty:
            elapsed = (time.time() - start_time) * 1000
            return ToolResult.failure("Both data_a and data_b must be provided", execution_time_ms=elapsed)

        try:
            import matplotlib
            matplotlib.use("Agg")
            import matplotlib.pyplot as plt
            import numpy as np

            if key_column and key_column in df_a.columns and key_column in df_b.columns:
                merged = df_a.merge(df_b, on=key_column, how="outer", suffixes=("_a", "_b"))
                key_vals = merged[key_column].astype(str)
            else:
                merged = pd.concat([df_a.reset_index(drop=True), df_b.reset_index(drop=True)], axis=1)
                key_vals = merged.index.astype(str)

            a_cols = [c for c in merged.columns if c.endswith("_a") and not c.endswith("_b")]
            b_cols = [c for c in merged.columns if c.endswith("_b") and not c.endswith("_a")]

            paired = []
            if compare_columns:
                paired = [(c, f"{c}_b") for c in compare_columns if f"{c}_b" in merged.columns]
            else:
                for ca in a_cols:
                    base = ca[:-2] if ca.endswith("_a") else ca
                    cb = f"{base}_b"
                    if cb in merged.columns:
                        paired.append((ca, cb))

            if not paired:
                common = set(df_a.select_dtypes(include=["number"]).columns) & set(df_b.select_dtypes(include=["number"]).columns)
                for col in common:
                    paired.append((col, col))

            comparison_rows = []
            summary = {"rows_added": 0, "rows_removed": 0, "rows_changed": 0, "changes": []}

            for idx, row in merged.iterrows():
                for ca, cb in paired:
                    va = row.get(ca, row.get(ca.replace("_a", ""), None))
                    vb = row.get(cb, row.get(cb.replace("_b", ""), None))
                    if va is None and vb is None:
                        continue
                    if va is None:
                        summary["rows_added"] += 1
                        comparison_rows.append({
                            "key": key_vals[idx] if idx < len(key_vals) else str(idx),
                            "column": ca.replace("_a", ""),
                            "value_a": "N/A", "value_b": vb,
                            "diff": vb, "pct_change": None,
                            "direction": "added",
                        })
                    elif vb is None:
                        summary["rows_removed"] += 1
                        comparison_rows.append({
                            "key": key_vals[idx] if idx < len(key_vals) else str(idx),
                            "column": ca.replace("_a", ""),
                            "value_a": va, "value_b": "N/A",
                            "diff": -va, "pct_change": None,
                            "direction": "removed",
                        })
                    elif isinstance(va, (int, float)) and isinstance(vb, (int, float)):
                        diff = vb - va
                        pct = ((vb - va) / abs(va) * 100) if va != 0 else None
                        summary["rows_changed"] += 1
                        direction = "increase" if diff > 0 else "decrease" if diff < 0 else "unchanged"
                        if diff != 0:
                            summary["changes"].append({"key": key_vals[idx], "column": ca, "diff": diff, "pct": pct})
                        comparison_rows.append({
                            "key": key_vals[idx] if idx < len(key_vals) else str(idx),
                            "column": ca.replace("_a", ""),
                            "value_a": round(va, 2), "value_b": round(vb, 2),
                            "diff": round(diff, 2), "pct_change": round(pct, 2) if pct is not None else None,
                            "direction": direction,
                        })

            fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 5))

            if comparison_rows:
                comp_df = pd.DataFrame(comparison_rows)
                plot_data = comp_df[comp_df["direction"].isin(["increase", "decrease"])].head(15)
                if not plot_data.empty:
                    labels = [f"{r['key']}\n{r['column']}" for _, r in plot_data.iterrows()]
                    diffs = plot_data["diff"].values
                    bar_colors = ["#27AE60" if d >= 0 else "#E74C3C" for d in diffs]
                    ax1.barh(range(len(diffs)), diffs, color=bar_colors, edgecolor="white")
                    ax1.set_yticks(range(len(diffs)))
                    ax1.set_yticklabels(labels, fontsize=8)
                    ax1.axvline(0, color="black", linewidth=0.8)
                    ax1.set_title("Value Changes", fontsize=12, fontweight="bold")
                    ax1.set_xlabel("Difference")
                    ax1.grid(axis="x", alpha=0.3)
                    ax1.spines["top"].set_visible(False)
                    ax1.spines["right"].set_visible(False)
                else:
                    ax1.text(0.5, 0.5, "No significant changes", ha="center", va="center", transform=ax1.transAxes)
            else:
                ax1.text(0.5, 0.5, "No comparison data", ha="center", va="center", transform=ax1.transAxes)

            labels_pie, sizes_pie, colors_pie = [], [], []
            if summary["rows_added"] > 0:
                labels_pie.append("Added"); sizes_pie.append(summary["rows_added"]); colors_pie.append("#2ECC71")
            if summary["rows_removed"] > 0:
                labels_pie.append("Removed"); sizes_pie.append(summary["rows_removed"]); colors_pie.append("#E74C3C")
            if summary["rows_changed"] > 0:
                pos = len([c for c in summary["changes"] if c["diff"] > 0])
                neg = len([c for c in summary["changes"] if c["diff"] < 0])
                if pos > 0:
                    labels_pie.append("Increased"); sizes_pie.append(pos); colors_pie.append("#27AE60")
                if neg > 0:
                    labels_pie.append("Decreased"); sizes_pie.append(neg); colors_pie.append("#E74C3C")

            if sizes_pie:
                ax2.pie(sizes_pie, labels=labels_pie, autopct="%1.1f%%",
                        colors=colors_pie, startangle=90,
                        wedgeprops={"edgecolor": "white", "linewidth": 1.5})
            else:
                ax2.text(0.5, 0.5, "No changes", ha="center", va="center", transform=ax2.transAxes)
            ax2.set_title("Change Summary", fontsize=12, fontweight="bold")

            fig.suptitle(title, fontsize=14, fontweight="bold", y=1.02, color="#1B3A5C")
            plt.tight_layout()

            filename = self._safe_filename(title, ".png")
            filepath = self._reports_dir / filename
            fig.savefig(filepath, dpi=300, bbox_inches="tight", facecolor="white")
            plt.close(fig)
            self._record_report(filepath)

            # HTML comparison report
            row_html = ""
            for row in comparison_rows[:100]:
                cls = row["direction"]
                pct_display = f'{row["pct_change"]:.2f}%' if row["pct_change"] is not None else "N/A"
                row_html += f"<tr><td>{row['key']}</td><td>{row['column']}</td><td>{row['value_a']}</td><td>{row['value_b']}</td><td class='{cls}'>{row['diff']}</td><td class='{cls}'>{pct_display}</td></tr>"

            timestamp = datetime.now().strftime("%B %d, %Y %H:%M")
            html = f"""<!DOCTYPE html>
<html lang="en">
<head><meta charset="UTF-8"><title>{title}</title>
<style>
body {{ font-family:'Segoe UI',sans-serif; background:#f5f7fa; padding:20px; }}
.container {{ max-width:1200px; margin:0 auto; }}
h1 {{ color:#1B3A5C; }}
.summary {{ display:flex; gap:16px; margin:20px 0; }}
.summary-card {{ background:white; border-radius:8px; padding:16px; flex:1; box-shadow:0 2px 8px rgba(0,0,0,0.06); }}
.summary-card .num {{ font-size:28px; font-weight:700; }}
.summary-card .lbl {{ font-size:12px; color:#888; text-transform:uppercase; }}
.positive {{ color:#27AE60; }} .negative {{ color:#E74C3C; }} .neutral {{ color:#888; }}
img {{ max-width:100%; margin:20px 0; border-radius:8px; box-shadow:0 2px 12px rgba(0,0,0,0.1); }}
table {{ width:100%; border-collapse:collapse; background:white; border-radius:8px; overflow:hidden; box-shadow:0 2px 8px rgba(0,0,0,0.06); }}
th {{ background:#1B3A5C; color:white; padding:10px; text-align:left; }}
td {{ padding:8px 10px; border-bottom:1px solid #eee; font-size:13px; }}
tr:nth-child(even) {{ background:#f8fafc; }}
.increase {{ color:#27AE60; font-weight:600; }}
.decrease {{ color:#E74C3C; font-weight:600; }}
</style></head>
<body>
<div class="container">
    <h1>{title}</h1>
    <p>Generated {timestamp}</p>
    <div class="summary">
        <div class="summary-card"><div class="num positive">{summary['rows_added']}</div><div class="lbl">Added</div></div>
        <div class="summary-card"><div class="num negative">{summary['rows_removed']}</div><div class="lbl">Removed</div></div>
        <div class="summary-card"><div class="num">{summary['rows_changed']}</div><div class="lbl">Changed</div></div>
        <div class="summary-card"><div class="num">{len(summary['changes'])}</div><div class="lbl">Significant</div></div>
    </div>
    <img src="{filename}" alt="Comparison Chart">
    <h2>Detailed Comparison</h2>
    <table>
        <thead><tr><th>Key</th><th>Column</th><th>Value A</th><th>Value B</th><th>Difference</th><th>% Change</th></tr></thead>
        <tbody>{row_html}</tbody>
    </table>
</div></body></html>"""

            html_filename = self._safe_filename(f"{title}_comparison", ".html")
            html_path = self._reports_dir / html_filename
            html_path.write_text(html, encoding="utf-8")
            self._record_report(html_path)

            elapsed = (time.time() - start_time) * 1000
            return ToolResult.success(
                f"Comparison report generated: {filename}, {html_filename}",
                data={
                    "chart": {"path": str(filepath), "filename": filename},
                    "html": {"path": str(html_path), "filename": html_filename},
                    "summary": summary,
                    "comparisons": len(comparison_rows),
                    "format": "png+html",
                },
                execution_time_ms=elapsed,
            )
        except ImportError as e:
            elapsed = (time.time() - start_time) * 1000
            return ToolResult.failure(f"Comparison requires matplotlib: {e}", execution_time_ms=elapsed)
        except Exception as e:
            plt.close("all")
            elapsed = (time.time() - start_time) * 1000
            return ToolResult.error_result(f"Comparison failed: {e}", error=str(e), execution_time_ms=elapsed)
