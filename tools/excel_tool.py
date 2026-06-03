# MIT License
#
# Copyright (c) 2026 Aryan Chavan
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in all
# copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

"""
ArynoxTech AI Agent - Production-Grade Excel Business Tool
===========================================================
Comprehensive Excel/spreadsheet operations for business data handling:
read, create, modify, analyze, gst_calc, inventory_report, formula,
chart, pivot_table, compare, template, merge_workbooks.
"""

import io
import logging
import math
import os
import re
import time
import traceback
from collections import defaultdict
from copy import copy
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

import pandas as pd

from openpyxl import load_workbook, Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.comments import Comment
from openpyxl.formatting.rule import (
    CellIsRule,
    ColorScaleRule,
    DataBarRule,
    IconSetRule,
)
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from tools.base_tool import BaseTool, ToolResult
from config.settings import BASE_DIR, DIRS, SECURITY_CONFIG

logger = logging.getLogger(__name__)


class ExcelTool(BaseTool):
    """
    Production-grade Excel business tool for comprehensive spreadsheet operations.

    Provides 12 action categories:
    - read, create, modify, analyze, gst_calc, inventory_report
    - formula, chart, pivot_table, compare, template, merge_workbooks
    """

    name: str = "excel_tool"
    description: str = (
        "Comprehensive Excel/spreadsheet tool: read, create, modify, analyze, "
        "GST calculate, inventory report, formulas, charts, pivot tables, "
        "file comparison, business templates, merge workbooks."
    )
    version: str = "1.0.0"

    def __init__(self) -> None:
        super().__init__()
        self.config = {}
        self._max_rows = 100000
        self._max_columns = 50
        self._allowed_extensions = [".xlsx", ".xls", ".xlsm", ".csv"]
        self._dataframes: Dict[str, pd.DataFrame] = {}

        self._style_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )
        self._header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self._header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        self._header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self._data_font = Font(size=10, name="Calibri")
        self._data_alignment = Alignment(vertical="center")
        self._thin_side = Side(style="thin", color="D9D9D9")

        self.supported_actions = {
            "read": self._read,
            "create": self._create,
            "modify": self._modify,
            "analyze": self._analyze,
            "gst_calc": self._gst_calc,
            "inventory_report": self._inventory_report,
            "formula": self._formula,
            "chart": self._chart,
            "pivot_table": self._pivot_table,
            "compare": self._compare,
            "template": self._template,
            "merge_workbooks": self._merge_workbooks,
        }

    # ── Dispatcher ──────────────────────────────────────────────────────────────

    async def execute(self, **kwargs: Any) -> ToolResult:
        start_time = time.time()
        action = kwargs.get("action", "read")

        try:
            handler = self.supported_actions.get(action)
            if handler is None:
                return ToolResult.failure(
                    f"Unknown action: {action}. Available: {list(self.supported_actions.keys())}",
                    execution_time_ms=(time.time() - start_time) * 1000,
                )
            return await handler(kwargs, start_time)
        except Exception as e:
            self.logger.exception(f"ExcelTool error: {e}")
            return ToolResult.error_result(
                f"Excel operation failed: {e}",
                error=traceback.format_exc(),
                execution_time_ms=(time.time() - start_time) * 1000,
            )

    # ── Helpers ─────────────────────────────────────────────────────────────────

    def _validate_path(self, file_path: str) -> Path:
        path = Path(file_path).resolve()
        if not path.parent.exists():
            path.parent.mkdir(parents=True, exist_ok=True)
        ext = path.suffix.lower()
        if ext not in self._allowed_extensions:
            raise ValueError(
                f"Invalid extension '{ext}'. Allowed: {self._allowed_extensions}"
            )
        return path

    def _resolve_data(
        self, kwargs: Dict, as_frame: bool = True
    ) -> Union[pd.DataFrame, List[Dict], None]:
        data = kwargs.get("data")
        file_path = kwargs.get("file_path", "")
        sheet_name = kwargs.get("sheet_name")

        if data is not None:
            if isinstance(data, pd.DataFrame):
                return data if as_frame else data.to_dict(orient="records")
            if isinstance(data, list):
                df = pd.DataFrame(data)
                return df if as_frame else data
            if isinstance(data, dict):
                df = pd.DataFrame([data])
                return df if as_frame else [data]
            return data

        if file_path:
            path = Path(file_path)
            if not path.exists():
                return None
            ext = path.suffix.lower()
            if ext == ".csv":
                df = pd.read_csv(file_path)
            else:
                df = pd.read_excel(file_path, sheet_name=sheet_name or 0)
            return df if as_frame else df.to_dict(orient="records")

        return None

    def _apply_header_style(self, ws, num_cols: int, row: int = 1) -> None:
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = self._header_font
            cell.fill = self._header_fill
            cell.alignment = self._header_alignment
            cell.border = Border(
                left=Side(style="thin", color="4472C4"),
                right=Side(style="thin", color="4472C4"),
                top=Side(style="thin", color="4472C4"),
                bottom=Side(style="medium", color="4472C4"),
            )

    def _apply_data_border(self, ws, max_row: int, max_col: int, start_row: int = 2) -> None:
        for r in range(start_row, max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = self._style_border
                cell.font = self._data_font
                cell.alignment = self._data_alignment

    def _auto_column_width(self, ws, max_col: int, max_row: int, max_width: int = 50) -> None:
        for col in range(1, max_col + 1):
            lengths = []
            for row in range(1, max_row + 1):
                val = ws.cell(row=row, column=col).value
                if val is not None:
                    lengths.append(len(str(val)))
            best = max(lengths) if lengths else 8
            ws.column_dimensions[get_column_letter(col)].width = min(best + 3, max_width)

    def _freeze_header(self, ws, cell: str = "A2") -> None:
        ws.freeze_panes = cell

    def _get_ws(self, wb, sheet_name: Optional[str] = None):
        if sheet_name and sheet_name in wb.sheetnames:
            return wb[sheet_name]
        return wb.active

    def _resolve_range(
        self, ws, range_str: Optional[str] = None
    ) -> Tuple[int, int, int, int]:
        if range_str:
            min_col, min_row, max_col, max_row = range_boundaries(range_str)
            return min_row, min_col, max_row, max_col
        return 1, 1, ws.max_row or 1, ws.max_column or 1

    # ── 1. read ─────────────────────────────────────────────────────────────────

    async def _read(self, kwargs: Dict, start_time: float) -> ToolResult:
        file_path = self._validate_path(kwargs.get("file_path", ""))
        sheet_name: Optional[str] = kwargs.get("sheet_name")
        range_str: Optional[str] = kwargs.get("range")
        return_as = kwargs.get("return_as", "dicts")

        if not file_path.exists():
            return ToolResult.failure(
                f"File not found: {file_path}",
                execution_time_ms=(time.time() - start_time) * 1000,
            )

        try:
            ext = file_path.suffix.lower()
            if ext == ".csv":
                df = pd.read_csv(file_path)
                sheets_data = {"Sheet1": df}
            else:
                wb = load_workbook(file_path, data_only=True, read_only=True)
                sheet_names = wb.sheetnames

                if sheet_name:
                    if sheet_name not in sheet_names:
                        wb.close()
                        return ToolResult.failure(
                            f"Sheet '{sheet_name}' not found. Available: {sheet_names}",
                            execution_time_ms=(time.time() - start_time) * 1000,
                        )
                    sheet_names = [sheet_name]

                sheets_data = {}
                for sn in sheet_names:
                    ws = wb[sn]
                    rows = list(ws.iter_rows(values_only=True))
                    if not rows:
                        sheets_data[sn] = pd.DataFrame()
                        continue
                    headers = [str(h) if h is not None else f"Column{i}" for i, h in enumerate(rows[0])]
                    data_rows = []
                    for row in rows[1:]:
                        padded = list(row) + [None] * (len(headers) - len(row))
                        data_rows.append(padded[: len(headers)])

                    df = pd.DataFrame(data_rows, columns=headers)
                    df = df.map(lambda x: x.strip() if isinstance(x, str) else x)

                    if range_str:
                        min_row, min_col, max_row, max_col = self._resolve_range(
                            None, range_str
                        )
                        df = df.iloc[
                            max(0, min_row - 2) : max_row - 1,
                            max(0, min_col - 1) : max_col,
                        ]
                        df.columns = [f"Column{i}" for i in range(df.shape[1])]

                    self._dataframes[f"{sn}"] = df
                    sheets_data[sn] = df
                wb.close()

            result = {}
            for sn, df in sheets_data.items():
                data = (
                    df.to_dict(orient="records")
                    if return_as == "dicts"
                    else df.to_dict(orient="list")
                )
                result[sn] = {
                    "columns": list(df.columns),
                    "rows": len(df),
                    "data": data,
                    "dtypes": {c: str(dt) for c, dt in df.dtypes.items()},
                }

            total_rows = sum(v["rows"] for v in result.values())
            elapsed = (time.time() - start_time) * 1000
            return ToolResult.success(
                f"Read {len(result)} sheet(s), {total_rows} total rows from {file_path.name}",
                data=result,
                execution_time_ms=elapsed,
            )
        except Exception as e:
            return ToolResult.error_result(
                f"Failed to read {file_path.name}: {e}",
                error=traceback.format_exc(),
                execution_time_ms=(time.time() - start_time) * 1000,
            )

    # ── 2. create ───────────────────────────────────────────────────────────────

    async def _create(self, kwargs: Dict, start_time: float) -> ToolResult:
        file_path = self._validate_path(kwargs.get("file_path", "output.xlsx"))
        data = kwargs.get("data")
        sheet_name = kwargs.get("sheet_name", "Sheet1")
        multi_sheets: Optional[List[Dict]] = kwargs.get("sheets")
        data_validations: Optional[List[Dict]] = kwargs.get("data_validations")
        conditional_formats: Optional[List[Dict]] = kwargs.get("conditional_formats")
        named_ranges: Optional[List[Dict]] = kwargs.get("named_ranges")
        comments: Optional[List[Dict]] = kwargs.get("comments")
        freeze: bool = kwargs.get("freeze_panes", True)

        try:
            wb = Workbook()

            if multi_sheets:
                for i, sheet_conf in enumerate(multi_sheets):
                    sn = sheet_conf.get("name", f"Sheet{i + 1}")
                    sd = sheet_conf.get("data", [])
                    if i == 0:
                        ws = wb.active
                        ws.title = sn
                    else:
                        ws = wb.create_sheet(title=sn)

                    df = pd.DataFrame(sd) if isinstance(sd, list) else sd
                    self._write_df_to_ws(ws, df)
                    self._dataframes[sn] = df

                    if freeze:
                        self._freeze_header(ws)

                    if data_validations:
                        self._apply_data_validations(ws, data_validations)
                    if conditional_formats:
                        self._apply_conditional_formats(ws, conditional_formats)
                    if named_ranges:
                        self._apply_named_ranges(wb, ws, named_ranges)
                    if comments:
                        self._apply_comments(ws, comments, sheet_name=sn)
            else:
                ws = wb.active
                ws.title = sheet_name

                if isinstance(data, pd.DataFrame):
                    df = data
                elif isinstance(data, list):
                    df = pd.DataFrame(data)
                elif isinstance(data, dict):
                    df = pd.DataFrame([data])
                else:
                    return ToolResult.failure(
                        "Unsupported data format. Use list of dicts or DataFrame.",
                        execution_time_ms=(time.time() - start_time) * 1000,
                    )

                self._write_df_to_ws(ws, df)
                self._dataframes[sheet_name] = df

                if freeze:
                    self._freeze_header(ws)

                if data_validations:
                    self._apply_data_validations(ws, data_validations)
                if conditional_formats:
                    self._apply_conditional_formats(ws, conditional_formats)
                if named_ranges:
                    self._apply_named_ranges(wb, ws, named_ranges)
                if comments:
                    self._apply_comments(ws, comments, sheet_name=sheet_name)

            wb.save(str(file_path))
            total_rows = sum(
                len(v) for v in self._dataframes.values()
            )
            elapsed = (time.time() - start_time) * 1000
            return ToolResult.success(
                f"Created {file_path.name} ({len(wb.sheetnames)} sheet(s), {total_rows} rows)",
                data={"path": str(file_path), "sheets": wb.sheetnames},
                execution_time_ms=elapsed,
            )
        except Exception as e:
            return ToolResult.error_result(
                f"Failed to create Excel file: {e}",
                error=traceback.format_exc(),
                execution_time_ms=(time.time() - start_time) * 1000,
            )

    def _write_df_to_ws(self, ws, df: pd.DataFrame) -> None:
        if df.empty:
            return
        df = df.fillna("").astype(str)
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx, value=str(col_name))
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            for col_idx, col_name in enumerate(df.columns, 1):
                ws.cell(row=row_idx, column=col_idx, value=row[col_name])

        self._apply_header_style(ws, len(df.columns))
        self._apply_data_border(ws, len(df) + 1, len(df.columns))
        self._auto_column_width(ws, len(df.columns), len(df) + 1)

    def _apply_data_validations(self, ws, validations: List[Dict]) -> None:
        for dv_conf in validations:
            dv_type = dv_conf.get("type", "list")
            formula1 = dv_conf.get("formula1", "")
            formula2 = dv_conf.get("formula2")
            sqref = dv_conf.get("sqref", ws.dimensions)
            allow_blank = dv_conf.get("allow_blank", True)
            show_input = dv_conf.get("show_input_message", True)
            show_error = dv_conf.get("show_error_alert", True)
            error_title = dv_conf.get("error_title", "Invalid Value")
            error_msg = dv_conf.get("error_message", "Please enter a valid value.")

            dv = DataValidation(
                type=dv_type,
                formula1=formula1,
                formula2=formula2,
                allow_blank=allow_blank,
                showInputMessage=show_input,
                showErrorMessage=show_error,
                errorTitle=error_title,
                error=error_msg,
            )
            dv.sqref = sqref
            ws.add_data_validation(dv)

    def _apply_conditional_formats(self, ws, formats: List[Dict]) -> None:
        for cf in formats:
            cf_type = cf.get("type", "color_scale")
            sqref = cf.get("sqref", ws.dimensions)

            if cf_type == "color_scale":
                colors = cf.get("colors", ["F8696B", "FFEB84", "63BE7B"])
                rule = ColorScaleRule(
                    start_type="min", start_color=colors[0],
                    mid_type="percentile", mid_value=50, mid_color=colors[1] if len(colors) >= 3 else None,
                    end_type="max", end_color=colors[-1],
                )
            elif cf_type == "data_bar":
                color = cf.get("color", "5A8AC6")
                rule = DataBarRule(
                    start_type="min", end_type="max",
                    color=color, showValue=True,
                )
            elif cf_type == "icon_set":
                icon_style = cf.get("icon_style", "3TrafficLights1")
                rule = IconSetRule(icon_style=icon_style)
            elif cf_type == "cell_is":
                operator = cf.get("operator", "greaterThan")
                formula = cf.get("formula", ["0"])
                fill_color = cf.get("fill", "FF0000")
                font_color = cf.get("font_color", "FFFFFF")
                rule = CellIsRule(
                    operator=operator,
                    formula=formula,
                    fill=PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid"),
                    font=Font(color=font_color),
                )
            else:
                continue

            ws.conditional_formatting.add(sqref, rule)

    def _apply_named_ranges(self, wb, ws, named_ranges: List[Dict]) -> None:
        for nr in named_ranges:
            name = nr.get("name", "Range")
            ref = nr.get("ref", ws.dimensions)
            scope = nr.get("scope", None)
            from openpyxl.workbook.defined_name import DefinedName
            dn = DefinedName(name, attr_text=ref)
            if scope:
                dn.localSheetId = wb.sheetnames.index(ws.title)
            wb.defined_names.add(dn)

    def _apply_comments(self, ws, comments: List[Dict], sheet_name: str = "Sheet1") -> None:
        for cm in comments:
            cell_ref = cm.get("cell", "A1")
            text = cm.get("text", "")
            author = cm.get("author", "ArynoxTech AI Agent")
            ws[cell_ref].comment = Comment(text, author)

    # ── 3. modify ──────────────────────────────────────────────────────────────

    async def _modify(self, kwargs: Dict, start_time: float) -> ToolResult:
        file_path = self._validate_path(kwargs.get("file_path", ""))
        if not file_path.exists():
            return ToolResult.failure(
                f"File not found: {file_path}",
                execution_time_ms=(time.time() - start_time) * 1000,
            )

        try:
            wb = load_workbook(file_path)
            sheet_name: Optional[str] = kwargs.get("sheet_name")
            ws = self._get_ws(wb, sheet_name)
            mod_type = kwargs.get("mod_type", "update_cells")

            if mod_type == "update_cells":
                updates = kwargs.get("updates", {})
                for cell_ref, value in updates.items():
                    ws[cell_ref] = value

            elif mod_type == "update_range":
                data = kwargs.get("data", [])
                start_cell = kwargs.get("start_cell", "A1")
                start_col = ord(start_cell[0].upper()) - ord("A") + 1
                start_row = int(start_cell[1:]) if start_cell[1:].isdigit() else 1
                for r_idx, row_data in enumerate(data):
                    for c_idx, val in enumerate(row_data if isinstance(row_data, list) else [row_data]):
                        ws.cell(row=start_row + r_idx, column=start_col + c_idx, value=val)

            elif mod_type == "insert_rows":
                row_idx = kwargs.get("row", 1)
                count = kwargs.get("count", 1)
                ws.insert_rows(row_idx, count)

            elif mod_type == "delete_rows":
                row_idx = kwargs.get("row", 1)
                count = kwargs.get("count", 1)
                ws.delete_rows(row_idx, count)

            elif mod_type == "insert_cols":
                col_idx = kwargs.get("col", 1)
                count = kwargs.get("count", 1)
                ws.insert_cols(col_idx, count)

            elif mod_type == "delete_cols":
                col_idx = kwargs.get("col", 1)
                count = kwargs.get("count", 1)
                ws.delete_cols(col_idx, count)

            elif mod_type == "add_sheet":
                new_name = kwargs.get("sheet_name", "NewSheet")
                wb.create_sheet(title=new_name)

            elif mod_type == "rename_sheet":
                old = kwargs.get("old_name", "")
                new = kwargs.get("new_name", "")
                if old in wb.sheetnames:
                    wb[old].title = new

            elif mod_type == "delete_sheet":
                target = kwargs.get("sheet_name", "")
                if target in wb.sheetnames and len(wb.sheetnames) > 1:
                    del wb[target]

            elif mod_type == "copy_sheet":
                source = kwargs.get("source_sheet", "")
                target = kwargs.get("target_sheet", f"{source}_copy")
                if source in wb.sheetnames:
                    src_ws = wb[source]
                    new_ws = wb.copy_worksheet(src_ws)
                    new_ws.title = target

            elif mod_type == "move_sheet":
                sheet = kwargs.get("sheet_name", "")
                position = kwargs.get("position", 0)
                if sheet in wb.sheetnames:
                    idx = wb.sheetnames.index(sheet)
                    wb.move_sheet(sheet, offset=position - idx)

            elif mod_type == "find_replace":
                find = kwargs.get("find", "")
                replace = kwargs.get("replace", "")
                scope = kwargs.get("scope", "sheet")
                if scope == "workbook":
                    for sn in wb.sheetnames:
                        s_ws = wb[sn]
                        self._find_replace_in_ws(s_ws, find, replace)
                else:
                    self._find_replace_in_ws(ws, find, replace)

            elif mod_type == "merge_cells":
                merge_range = kwargs.get("range", "")
                if merge_range:
                    ws.merge_cells(merge_range)

            elif mod_type == "unmerge_cells":
                unmerge_range = kwargs.get("range", "")
                if unmerge_range:
                    ws.unmerge_cells(unmerge_range)

            elif mod_type == "format_range":
                fmt_range = kwargs.get("range", ws.dimensions)
                fmt_config = kwargs.get("format", {})
                self._apply_format_to_range(ws, fmt_range, fmt_config)

            else:
                wb.close()
                return ToolResult.failure(
                    f"Unknown mod_type: {mod_type}",
                    execution_time_ms=(time.time() - start_time) * 1000,
                )

            wb.save(str(file_path))
            wb.close()
            elapsed = (time.time() - start_time) * 1000
            return ToolResult.success(
                f"Modified {file_path.name} ({mod_type})",
                data={"path": str(file_path), "mod_type": mod_type},
                execution_time_ms=elapsed,
            )
        except Exception as e:
            return ToolResult.error_result(
                f"Modify failed: {e}",
                error=traceback.format_exc(),
                execution_time_ms=(time.time() - start_time) * 1000,
            )

    def _find_replace_in_ws(self, ws, find: str, replace: str) -> int:
        count = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and isinstance(cell.value, str) and find in cell.value:
                    cell.value = cell.value.replace(find, replace)
                    count += 1
        return count

    def _apply_format_to_range(self, ws, range_str: str, fmt: Dict) -> None:
        min_col, min_row, max_col, max_row = range_boundaries(range_str)
        font_kw = {k: v for k, v in fmt.items() if k in ("bold", "italic", "size", "color", "name")}
        fill_color = fmt.get("fill_color")
        align_h = fmt.get("align_horizontal")
        align_v = fmt.get("align_vertical")
        number_format = fmt.get("number_format")
        border_style = fmt.get("border_style", "thin")
        border_color = fmt.get("border_color", "D9D9D9")

        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                cell = ws.cell(row=r, column=c)
                if font_kw:
                    cell.font = Font(**font_kw)
                if fill_color:
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                if align_h or align_v:
                    cell.alignment = Alignment(horizontal=align_h, vertical=align_v, wrap_text=fmt.get("wrap_text", False))
                if number_format:
                    cell.number_format = number_format
                if border_style != "none":
                    side = Side(style=border_style, color=border_color)
                    cell.border = Border(left=side, right=side, top=side, bottom=side)

    # ── 4. analyze ─────────────────────────────────────────────────────────────

    async def _analyze(self, kwargs: Dict, start_time: float) -> ToolResult:
        data = self._resolve_data(kwargs, as_frame=True)
        if data is None:
            return ToolResult.failure(
                "Provide data, file_path, or sheet_name for analysis",
                execution_time_ms=(time.time() - start_time) * 1000,
            )

        try:
            df = data.copy()
            analysis: Dict[str, Any] = {
                "overview": {
                    "rows": len(df),
                    "columns": len(df.columns),
                    "column_names": list(df.columns),
                    "dtypes": {c: str(dt) for c, dt in df.dtypes.items()},
                    "memory_usage_kb": round(df.memory_usage(deep=True).sum() / 1024, 2),
                    "duplicate_rows": int(df.duplicated().sum()),
                    "missing_cells": int(df.isnull().sum().sum()),
                    "missing_pct": round(df.isnull().sum().sum() / (len(df) * len(df.columns)) * 100, 2),
                },
            }

            numeric_cols = df.select_dtypes(include=["number"]).columns.tolist()
            text_cols = df.select_dtypes(include=["object", "category"]).columns.tolist()
            datetime_cols = df.select_dtypes(include=["datetime"]).columns.tolist()

            if numeric_cols:
                desc = df[numeric_cols].describe(percentiles=[0.05, 0.25, 0.5, 0.75, 0.95]).round(4)
                desc.index = desc.index.astype(str)
                analysis["numeric_summary"] = desc.to_dict()
                analysis["numeric_columns"] = numeric_cols
                analysis["skewness"] = df[numeric_cols].skew().round(4).to_dict()
                analysis["kurtosis"] = df[numeric_cols].kurtosis().round(4).to_dict()
                analysis["correlation_matrix"] = (
                    df[numeric_cols].corr().round(4).to_dict()
                    if len(numeric_cols) > 1
                    else {}
                )

            if text_cols:
                analysis["text_columns"] = text_cols
                analysis["value_counts"] = {}
                for col in text_cols:
                    vc = df[col].value_counts(dropna=False).head(10)
                    analysis["value_counts"][col] = {
                        "unique": int(df[col].nunique()),
                        "top_values": vc.to_dict(),
                        "missing": int(df[col].isnull().sum()),
                    }

            analysis["missing_summary"] = {
                "total": int(df.isnull().sum().sum()),
                "per_column": df.isnull().sum().to_dict(),
                "per_column_pct": df.isnull().mean().mul(100).round(2).to_dict(),
            }

            analysis["summary_text"] = self._generate_analysis_text(df, analysis)

            elapsed = (time.time() - start_time) * 1000
            return ToolResult.success(
                f"Analysis complete: {len(df)} rows, {len(df.columns)} columns",
                data=analysis,
                execution_time_ms=elapsed,
            )
        except Exception as e:
            return ToolResult.error_result(
                f"Analysis failed: {e}",
                error=traceback.format_exc(),
                execution_time_ms=(time.time() - start_time) * 1000,
            )

    def _generate_analysis_text(self, df: pd.DataFrame, analysis: Dict) -> str:
        lines = []
        o = analysis["overview"]
        lines.append(f"Dataset: {o['rows']} rows x {o['columns']} columns")
        lines.append(f"Columns: {', '.join(o['column_names'])}")
        lines.append(f"Missing cells: {o['missing_cells']} ({o['missing_pct']}%)")
        lines.append(f"Duplicate rows: {o['duplicate_rows']}")

        nc = analysis.get("numeric_columns", [])
        if nc:
            lines.append(f"\nNumeric columns ({len(nc)}): {', '.join(nc)}")
            ns = analysis.get("numeric_summary", {})
            if ns:
                for col in nc[:5]:
                    stats = ns.get(col, {})
                    lines.append(
                        f"  {col}: mean={stats.get('mean','')}, "
                        f"min={stats.get('min','')}, max={stats.get('max','')}"
                    )

        tc = analysis.get("text_columns", [])
        if tc:
            lines.append(f"\nText columns ({len(tc)}): {', '.join(tc)}")
            for col in tc[:5]:
                vc = analysis.get("value_counts", {}).get(col, {})
                lines.append(f"  {col}: {vc.get('unique', 0)} unique values")

        return "\n".join(lines)

    # ── 5. gst_calc ────────────────────────────────────────────────────────────

    async def _gst_calc(self, kwargs: Dict, start_time: float) -> ToolResult:
        line_items: List[Dict] = kwargs.get("line_items", [])
        default_gst_rate = kwargs.get("default_gst_rate", 18.0)
        output_path: Optional[str] = kwargs.get("output_path")
        business_name = kwargs.get("business_name", "ArynoxTech")
        gstin = kwargs.get("gstin", "")
        invoice_no = kwargs.get("invoice_no", f"INV-{datetime.now().strftime('%Y%m%d%H%M%S')}")
        invoice_date = kwargs.get("invoice_date", datetime.now().strftime("%d-%m-%Y"))
        buyer_details = kwargs.get("buyer_details", "")

        if not line_items and kwargs.get("data"):
            resolved = self._resolve_data(kwargs, as_frame=True)
            if resolved is not None:
                line_items = resolved.to_dict(orient="records")
            else:
                line_items = kwargs.get("data", [])

        if not line_items:
            single_amount = kwargs.get("amount")
            if single_amount is not None:
                line_items = [{"description": "Service", "amount": float(single_amount), "gst_rate": default_gst_rate}]
            else:
                return ToolResult.failure(
                    "Provide line_items (list of dicts) or amount for GST calculation",
                    execution_time_ms=(time.time() - start_time) * 1000,
                )

        try:
            total_taxable = 0.0
            total_cgst = 0.0
            total_sgst = 0.0
            total_igst = 0.0
            total_gst = 0.0
            grand_total = 0.0

            items_detail = []
            for i, item in enumerate(line_items, 1):
                desc = item.get("description", f"Item {i}")
                amount = float(item.get("amount", 0))
                gst_rate = float(item.get("gst_rate", default_gst_rate))
                hsn = item.get("hsn", f"HSN{i:04d}")
                quantity = int(item.get("quantity", 1))
                unit_price = float(item.get("unit_price", amount))

                taxable = amount / (1 + gst_rate / 100)
                gst_amt = amount - taxable
                cgst = gst_amt / 2
                sgst = gst_amt / 2
                igst = gst_amt if kwargs.get("gst_type", "regular") == "igst" else 0.0
                if kwargs.get("gst_type", "regular") != "igst":
                    cgst = gst_amt / 2
                    sgst = gst_amt / 2
                    igst = 0.0
                else:
                    cgst = 0.0
                    sgst = 0.0
                    igst = gst_amt

                item_total = taxable + gst_amt

                items_detail.append({
                    "sl_no": i,
                    "hsn": hsn,
                    "description": desc,
                    "quantity": quantity,
                    "unit_price": round(unit_price, 2),
                    "amount": round(amount, 2),
                    "taxable_value": round(taxable, 2),
                    "gst_rate": gst_rate,
                    "cgst": round(cgst, 2),
                    "sgst": round(sgst, 2),
                    "igst": round(igst, 2),
                    "total": round(item_total, 2),
                })

                total_taxable += taxable
                total_cgst += cgst
                total_sgst += sgst
                total_igst += igst
                total_gst += gst_amt
                grand_total += item_total

            invoice_data = {
                "invoice_no": invoice_no,
                "invoice_date": invoice_date,
                "business_name": business_name,
                "gstin": gstin,
                "buyer_details": buyer_details,
                "items": items_detail,
                "summary": {
                    "total_taxable_value": round(total_taxable, 2),
                    "total_cgst": round(total_cgst, 2),
                    "total_sgst": round(total_sgst, 2),
                    "total_igst": round(total_igst, 2),
                    "total_gst": round(total_gst, 2),
                    "grand_total": round(grand_total, 2),
                    "total_items": len(items_detail),
                },
            }

            if output_path:
                out_path = self._validate_path(output_path)
                self._generate_gst_excel(out_path, invoice_data)

            elapsed = (time.time() - start_time) * 1000
            return ToolResult.success(
                f"GST calc complete: {len(items_detail)} items, "
                f"Taxable=₹{total_taxable:,.2f}, GST=₹{total_gst:,.2f}, "
                f"Total=₹{grand_total:,.2f}",
                data=invoice_data,
                execution_time_ms=elapsed,
            )
        except Exception as e:
            return ToolResult.error_result(
                f"GST calculation failed: {e}",
                error=traceback.format_exc(),
                execution_time_ms=(time.time() - start_time) * 1000,
            )

    def _generate_gst_excel(self, path: Path, invoice: Dict) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "GST Invoice"

        inv = invoice["summary"]
        items = invoice["items"]

        title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        title_font = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
        sub_font = Font(bold=True, size=10, name="Calibri")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        ws.merge_cells("A1:H1")
        ws["A1"] = f"TAX INVOICE - {invoice['business_name']}"
        ws["A1"].font = title_font
        ws["A1"].fill = title_fill
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 35

        ws["A3"] = f"Invoice No: {invoice['invoice_no']}"
        ws["B3"] = f"Date: {invoice['invoice_date']}"
        ws["A3"].font = sub_font
        ws["B3"].font = sub_font
        if invoice["gstin"]:
            ws["A4"] = f"GSTIN: {invoice['gstin']}"
        if invoice["buyer_details"]:
            ws["B4"] = f"Buyer: {invoice['buyer_details']}"

        headers = ["Sl No", "HSN", "Description", "Qty", "Taxable Value", "CGST", "SGST", "Total"]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=6, column=ci, value=h)
            cell.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for ri, item in enumerate(items, 7):
            vals = [item["sl_no"], item["hsn"], item["description"], item["quantity"],
                    item["taxable_value"], item["cgst"], item["sgst"], item["total"]]
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.border = thin_border
                cell.font = Font(size=10, name="Calibri")
                cell.alignment = Alignment(horizontal="center", vertical="center")

        summary_row = 7 + len(items)
        ws.merge_cells(f"A{summary_row}:E{summary_row}")
        ws.cell(row=summary_row, column=1, value="Total").font = sub_font
        ws.cell(row=summary_row, column=6, value=inv["total_cgst"]).font = sub_font
        ws.cell(row=summary_row, column=7, value=inv["total_sgst"]).font = sub_font
        ws.cell(row=summary_row, column=8, value=inv["grand_total"]).font = sub_font
        for ci in range(1, 9):
            ws.cell(row=summary_row, column=ci).border = thin_border

        grand_row = summary_row + 1
        ws.merge_cells(f"A{grand_row}:E{grand_row}")
        ws.cell(row=grand_row, column=1, value=f"Grand Total (₹ {inv['grand_total']:,.2f})").font = Font(bold=True, size=12)
        ws.cell(row=grand_row, column=8, value=inv["grand_total"]).font = Font(bold=True, size=12)

        self._auto_column_width(ws, 8, grand_row)
        self._freeze_header(ws, "A7")
        wb.save(str(path))

    # ── 6. inventory_report ────────────────────────────────────────────────────

    async def _inventory_report(self, kwargs: Dict, start_time: float) -> ToolResult:
        data = self._resolve_data(kwargs, as_frame=True)
        if data is None:
            return ToolResult.failure(
                "Provide inventory data or file_path",
                execution_time_ms=(time.time() - start_time) * 1000,
            )

        try:
            df = data.copy()
            df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]

            required = {"item", "quantity"}
            missing = required - set(df.columns)
            if missing:
                return ToolResult.failure(
                    f"Missing required columns: {missing}. Have: {list(df.columns)}",
                    execution_time_ms=(time.time() - start_time) * 1000,
                )

            item_col = "item"
            qty_col = "quantity"
            price_col = "unit_price" if "unit_price" in df.columns else ("price" if "price" in df.columns else None)
            category_col = "category" if "category" in df.columns else ("department" if "department" in df.columns else None)
            reorder_col = "reorder_level" if "reorder_level" in df.columns else ("reorder" if "reorder" in df.columns else None)

            df[qty_col] = pd.to_numeric(df[qty_col], errors="coerce").fillna(0)
            if price_col:
                df[price_col] = pd.to_numeric(df[price_col], errors="coerce").fillna(0)
            if reorder_col:
                df[reorder_col] = pd.to_numeric(df[reorder_col], errors="coerce").fillna(0)

            df["total_value"] = df[qty_col] * (df[price_col] if price_col else 1)

            reorder_level = kwargs.get("reorder_threshold")
            if reorder_col:
                reorder_level = reorder_level or 0
                df["reorder_level_used"] = df[reorder_col]
            else:
                reorder_level = reorder_level or 10
                df["reorder_level_used"] = reorder_level

            def _stock_status(row):
                if row[qty_col] <= 0:
                    return "Out of Stock"
                if row[qty_col] < row["reorder_level_used"]:
                    return "Low"
                if row[qty_col] > row["reorder_level_used"] * 3:
                    return "Overstock"
                return "Normal"

            df["stock_status"] = df.apply(_stock_status, axis=1)
            df["reorder_suggested"] = df.apply(
                lambda r: max(0, int(r["reorder_level_used"] * 2 - r[qty_col]))
                if r[qty_col] < r["reorder_level_used"] else 0,
                axis=1,
            )

            detail = df.to_dict(orient="records")

            total_qty = int(df[qty_col].sum())
            total_value = float(df["total_value"].sum())
            item_count = len(df)
            low_stock = int((df["stock_status"] == "Low").sum())
            out_of_stock = int((df["stock_status"] == "Out of Stock").sum())
            overstock = int((df["stock_status"] == "Overstock").sum())
            normal = int((df["stock_status"] == "Normal").sum())

            summary = {
                "total_items": item_count,
                "total_quantity": total_qty,
                "total_value": round(total_value, 2),
                "stock_breakdown": {
                    "normal": normal,
                    "low": low_stock,
                    "out_of_stock": out_of_stock,
                    "overstock": overstock,
                },
                "reorder_suggestions": int(df["reorder_suggested"].sum()),
            }

            if category_col:
                cat_stats = df.groupby(category_col).agg(
                    item_count=(item_col, "count"),
                    total_qty=(qty_col, "sum"),
                    total_value=("total_value", "sum"),
                    low_stock=(qty_col, lambda x: int((x < df.loc[x.index, "reorder_level_used"]).sum())),
                ).reset_index()
                summary["category_breakdown"] = cat_stats.to_dict(orient="records")

            if reorder_col:
                flagged = df[df[qty_col] < df[reorder_col]].to_dict(orient="records")
                summary["items_below_reorder"] = flagged

            output_path = kwargs.get("output_path")
            if output_path:
                out = self._validate_path(output_path)
                self._generate_inventory_excel(out, df, summary)

            elapsed = (time.time() - start_time) * 1000
            return ToolResult.success(
                f"Inventory report: {item_count} items, {total_qty} qty, "
                f"₹{total_value:,.2f}, {low_stock + out_of_stock} items need reorder",
                data={"summary": summary, "detail": detail},
                execution_time_ms=elapsed,
            )
        except Exception as e:
            return ToolResult.error_result(
                f"Inventory report failed: {e}",
                error=traceback.format_exc(),
                execution_time_ms=(time.time() - start_time) * 1000,
            )

    def _generate_inventory_excel(self, path: Path, df: pd.DataFrame, summary: Dict) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory Report"

        self._write_df_to_ws(ws, df[["item", "quantity", "total_value", "stock_status", "reorder_suggested"]
                                     if all(c in df.columns for c in ["item", "quantity", "total_value", "stock_status", "reorder_suggested"])
                                     else df.columns.tolist()])

        low_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        out_fill = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
        over_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        status_col = list(df.columns).index("stock_status") + 1 if "stock_status" in df.columns else None
        if status_col:
            for row in range(2, ws.max_row + 1):
                val = ws.cell(row=row, column=status_col).value
                if val == "Low":
                    for c in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=c).fill = low_fill
                elif val == "Out of Stock":
                    for c in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=c).fill = out_fill
                elif val == "Overstock":
                    for c in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=c).fill = over_fill

        sr = ws.max_row + 2
        ws.cell(row=sr, column=1, value="Summary").font = Font(bold=True, size=12)
        for i, (k, v) in enumerate(summary.items(), sr + 1):
            if isinstance(v, dict):
                continue
            ws.cell(row=i, column=1, value=str(k).replace("_", " ").title())
            ws.cell(row=i, column=2, value=v)

        self._auto_column_width(ws, ws.max_column, ws.max_row)
        wb.save(str(path))

    # ── 7. formula ─────────────────────────────────────────────────────────────

    async def _formula(self, kwargs: Dict, start_time: float) -> ToolResult:
        file_path = self._validate_path(kwargs.get("file_path", ""))
        if not file_path.exists():
            return ToolResult.failure(
                f"File not found: {file_path}",
                execution_time_ms=(time.time() - start_time) * 1000,
            )

        try:
            wb = load_workbook(file_path)
            sheet_name: Optional[str] = kwargs.get("sheet_name")
            ws = self._get_ws(wb, sheet_name)

            formulas = kwargs.get("formulas", [])
            if not formulas:
                formula_type = kwargs.get("formula_type", "SUM")
                target_cell = kwargs.get("target_cell", "A1")
                source_range = kwargs.get("source_range", "")
                if not source_range and ws.max_row > 1:
                    source_range = f"A2:{get_column_letter(ws.max_column)}{ws.max_row}"

                formula_map = {
                    "SUM": f"=SUM({source_range})",
                    "AVERAGE": f"=AVERAGE({source_range})",
                    "COUNT": f"=COUNT({source_range})",
                    "MAX": f"=MAX({source_range})",
                    "MIN": f"=MIN({source_range})",
                    "IF": f'=IF({source_range},"Yes","No")',
                    "VLOOKUP": kwargs.get("vlookup_formula", f'=VLOOKUP({source_range})'),
                    "CONCATENATE": f'=CONCATENATE({source_range})',
                }
                formula_str = formula_map.get(formula_type.upper(), f"={formula_type}({source_range})")
                ws[target_cell] = formula_str
            else:
                for f in formulas:
                    cell_ref = f.get("cell", "A1")
                    formula_str = f.get("formula", "")
                    ws[cell_ref] = formula_str

            wb.save(str(file_path))
            wb.close()
            elapsed = (time.time() - start_time) * 1000
            return ToolResult.success(
                f"Formulas written to {file_path.name}",
                data={"path": str(file_path), "formulas_count": len(formulas) if formulas else 1},
                execution_time_ms=elapsed,
            )
        except Exception as e:
            return ToolResult.error_result(
                f"Formula write failed: {e}",
                error=traceback.format_exc(),
                execution_time_ms=(time.time() - start_time) * 1000,
            )

    # ── 8. chart ───────────────────────────────────────────────────────────────

    async def _chart(self, kwargs: Dict, start_time: float) -> ToolResult:
        file_path = self._validate_path(kwargs.get("file_path", ""))
        if not file_path.exists():
            return ToolResult.failure(
                f"File not found: {file_path}",
                execution_time_ms=(time.time() - start_time) * 1000,
            )

        try:
            wb = load_workbook(file_path)
            sheet_name: Optional[str] = kwargs.get("sheet_name")
            ws = self._get_ws(wb, sheet_name)

            chart_type = kwargs.get("chart_type", "bar").lower()
            data_range = kwargs.get("data_range", "")
            categories_range = kwargs.get("categories_range", "")
            title = kwargs.get("title", "Chart")
            x_axis = kwargs.get("x_axis", "")
            y_axis = kwargs.get("y_axis", "")
            placement = kwargs.get("placement", "E2")
            chart_colors = kwargs.get("colors", None)
            show_legend = kwargs.get("show_legend", True)
            show_data_labels = kwargs.get("show_data_labels", False)
            width = kwargs.get("width", 15)
            height = kwargs.get("height", 10)

            if not data_range:
                data_range = ws.dimensions

            data_ref = Reference(ws, range_string=data_range) if ":" in data_range else Reference(ws, min_col=1, min_row=1, max_col=ws.max_column, max_row=ws.max_row)
            cats_ref = None
            if categories_range and ":" in categories_range:
                cats_ref = Reference(ws, range_string=categories_range)
            elif categories_range:
                cats_ref = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

            chart_map = {
                "bar": BarChart,
                "column": BarChart,
                "line": LineChart,
                "pie": PieChart,
                "scatter": ScatterChart,
                "doughnut": PieChart,
            }

            chart_class = chart_map.get(chart_type, BarChart)
            chart = chart_class()

            if chart_type == "column":
                chart.type = "col"
            elif chart_type == "bar":
                chart.type = "bar"
            elif chart_type == "doughnut":
                from openpyxl.chart import DoughnutChart
                chart = DoughnutChart()

            if chart_type in ("pie", "doughnut"):
                chart.add_data(data_ref, titles_from_data=True)
                if cats_ref:
                    chart.set_categories(cats_ref)
            else:
                chart.add_data(data_ref, titles_from_data=True)
                if cats_ref:
                    chart.set_categories(cats_ref)

            chart.title = title
            chart.x_axis.title = x_axis
            chart.y_axis.title = y_axis
            chart.legend.position = "b" if show_legend else None
            chart.width = width
            chart.height = height

            if show_data_labels:
                chart.dataLabels = DataLabelList()
                chart.dataLabels.showVal = True

            if chart_colors:
                for i, series in enumerate(chart.series):
                    if i < len(chart_colors):
                        pt = DataPoint(idx=i)
                        pt.graphicalProperties.solidFill = chart_colors[i]
                        series.data_points.append(pt)

            ws.add_chart(chart, placement)
            wb.save(str(file_path))
            wb.close()
            elapsed = (time.time() - start_time) * 1000
            return ToolResult.success(
                f"Chart '{title}' ({chart_type}) added to {file_path.name}",
                data={"path": str(file_path), "chart_type": chart_type, "title": title},
                execution_time_ms=elapsed,
            )
        except Exception as e:
            return ToolResult.error_result(
                f"Chart creation failed: {e}",
                error=traceback.format_exc(),
                execution_time_ms=(time.time() - start_time) * 1000,
            )

    # ── 9. pivot_table ─────────────────────────────────────────────────────────

    async def _pivot_table(self, kwargs: Dict, start_time: float) -> ToolResult:
        data = self._resolve_data(kwargs, as_frame=True)
        if data is None:
            return ToolResult.failure(
                "Provide data or file_path for pivot table",
                execution_time_ms=(time.time() - start_time) * 1000,
            )

        try:
            df = data.copy()

            rows = kwargs.get("rows", [])
            if isinstance(rows, str):
                rows = [rows]
            columns = kwargs.get("columns", [])
            if isinstance(columns, str):
                columns = [columns]
            values = kwargs.get("values", [])
            if isinstance(values, str):
                values = [values]
            aggfunc = kwargs.get("aggfunc", "sum")
            fill_value = kwargs.get("fill_value", 0)
            margins = kwargs.get("margins", True)
            margins_name = kwargs.get("margins_name", "Grand Total")
            output_path = kwargs.get("output_path")

            if not rows or not values:
                return ToolResult.failure(
                    "pivot_table requires 'rows' (list of column names) and 'values' (list of column names)",
                    execution_time_ms=(time.time() - start_time) * 1000,
                )

            missing_cols = set(rows + columns + values) - set(df.columns)
            if missing_cols:
                return ToolResult.failure(
                    f"Columns not found in data: {missing_cols}",
                    execution_time_ms=(time.time() - start_time) * 1000,
                )

            pt = df.pivot_table(
                index=rows if rows else None,
                columns=columns if columns else None,
                values=values,
                aggfunc=aggfunc,
                fill_value=fill_value,
                margins=margins,
                margins_name=margins_name,
            ).reset_index()

            pt.columns = [str(c) if isinstance(c, tuple) else c for c in pt.columns]

            result = {
                "rows": rows,
                "columns": columns,
                "values": values,
                "aggfunc": aggfunc,
                "shape": list(pt.shape),
                "columns_list": list(pt.columns),
                "data": pt.head(1000).to_dict(orient="records"),
            }

            if output_path:
                out = self._validate_path(output_path)
                wb = Workbook()
                ws = wb.active
                ws.title = "Pivot Table"
                self._write_df_to_ws(ws, pt)
                wb.save(str(out))

            self._dataframes["pivot_result"] = pt

            elapsed = (time.time() - start_time) * 1000
            return ToolResult.success(
                f"Pivot table: {pt.shape[0]} rows, {pt.shape[1]} columns (agg={aggfunc})",
                data=result,
                execution_time_ms=elapsed,
            )
        except Exception as e:
            return ToolResult.error_result(
                f"Pivot table failed: {e}",
                error=traceback.format_exc(),
                execution_time_ms=(time.time() - start_time) * 1000,
            )

    # ── 10. compare ────────────────────────────────────────────────────────────

    async def _compare(self, kwargs: Dict, start_time: float) -> ToolResult:
        file1 = kwargs.get("file1", "")
        file2 = kwargs.get("file2", "")
        sheet1 = kwargs.get("sheet1")
        sheet2 = kwargs.get("sheet2")
        ignore_empty = kwargs.get("ignore_empty", True)
        output_path = kwargs.get("output_path")

        if not file1 or not file2:
            return ToolResult.failure(
                "compare requires file1 and file2 paths",
                execution_time_ms=(time.time() - start_time) * 1000,
            )

        try:
            path1 = self._validate_path(file1)
            path2 = self._validate_path(file2)

            if not path1.exists():
                return ToolResult.failure(f"File not found: {file1}")
            if not path2.exists():
                return ToolResult.failure(f"File not found: {file2}")

            wb1 = load_workbook(path1, data_only=True)
            wb2 = load_workbook(path2, data_only=True)

            if sheet1 and sheet1 not in wb1.sheetnames:
                wb1.close()
                wb2.close()
                return ToolResult.failure(f"Sheet '{sheet1}' not found in {path1.name}")
            if sheet2 and sheet2 not in wb2.sheetnames:
                wb1.close()
                wb2.close()
                return ToolResult.failure(f"Sheet '{sheet2}' not found in {path2.name}")

            ws1 = self._get_ws(wb1, sheet1) if sheet1 else wb1.active
            ws2 = self._get_ws(wb2, sheet2) if sheet2 else wb2.active

            diff_data = []
            max_row = max(ws1.max_row or 0, ws2.max_row or 0)
            max_col = max(ws1.max_column or 0, ws2.max_column or 0)
            changes = 0
            additions = 0
            deletions = 0

            for r in range(1, max_row + 1):
                for c in range(1, max_col + 1):
                    val1 = ws1.cell(row=r, column=c).value
                    val2 = ws2.cell(row=r, column=c).value

                    if ignore_empty and not val1 and not val2:
                        continue

                    cell_ref = f"{get_column_letter(c)}{r}"

                    if val1 is None and val2 is not None:
                        additions += 1
                        diff_data.append({
                            "cell": cell_ref,
                            "status": "added",
                            "old_value": "",
                            "new_value": str(val2),
                            "row": r,
                            "col": c,
                        })
                    elif val1 is not None and val2 is None:
                        deletions += 1
                        diff_data.append({
                            "cell": cell_ref,
                            "status": "deleted",
                            "old_value": str(val1),
                            "new_value": "",
                            "row": r,
                            "col": c,
                        })
                    elif str(val1) != str(val2):
                        changes += 1
                        diff_data.append({
                            "cell": cell_ref,
                            "status": "changed",
                            "old_value": str(val1),
                            "new_value": str(val2),
                            "row": r,
                            "col": c,
                        })

            total_diffs = len(diff_data)

            if output_path:
                out = self._validate_path(output_path)
                wb_out = Workbook()
                ws_out = wb_out.active
                ws_out.title = "Comparison Report"

                diff_headers = ["Cell", "Status", "Old Value", "New Value", "Row", "Column"]
                for ci, h in enumerate(diff_headers, 1):
                    ws_out.cell(row=1, column=ci, value=h)
                self._apply_header_style(ws_out, len(diff_headers))

                changed_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                added_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                deleted_fill = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")

                for ri, d in enumerate(diff_data, 2):
                    for ci, key in enumerate(["cell", "status", "old_value", "new_value", "row", "col"], 1):
                        cell = ws_out.cell(row=ri, column=ci, value=d[key])
                        cell.border = self._style_border
                    if d["status"] == "changed":
                        for ci in range(1, len(diff_headers) + 1):
                            ws_out.cell(row=ri, column=ci).fill = changed_fill
                    elif d["status"] == "added":
                        for ci in range(1, len(diff_headers) + 1):
                            ws_out.cell(row=ri, column=ci).fill = added_fill
                    elif d["status"] == "deleted":
                        for ci in range(1, len(diff_headers) + 1):
                            ws_out.cell(row=ri, column=ci).fill = deleted_fill

                sr = len(diff_data) + 3
                ws_out.cell(row=sr, column=1, value="Comparison Summary").font = Font(bold=True, size=12)
                ws_out.cell(row=sr + 1, column=1, value=f"File 1: {path1.name}")
                ws_out.cell(row=sr + 2, column=1, value=f"File 2: {path2.name}")
                ws_out.cell(row=sr + 3, column=1, value=f"Total Differences: {total_diffs}")
                ws_out.cell(row=sr + 4, column=1, value=f"Changes: {changes}")
                ws_out.cell(row=sr + 5, column=1, value=f"Additions: {additions}")
                ws_out.cell(row=sr + 6, column=1, value=f"Deletions: {deletions}")

                self._auto_column_width(ws_out, len(diff_headers), ws_out.max_row)
                wb_out.save(str(out))

            wb1.close()
            wb2.close()

            elapsed = (time.time() - start_time) * 1000
            return ToolResult.success(
                f"Comparison: {total_diffs} differences ({changes} changed, {additions} added, {deletions} deleted)",
                data={
                    "file1": str(path1),
                    "file2": str(path2),
                    "total_differences": total_diffs,
                    "changes": changes,
                    "additions": additions,
                    "deletions": deletions,
                    "differences": diff_data[:100],
                    "output_path": str(output_path) if output_path else None,
                },
                execution_time_ms=elapsed,
            )
        except Exception as e:
            return ToolResult.error_result(
                f"Comparison failed: {e}",
                error=traceback.format_exc(),
                execution_time_ms=(time.time() - start_time) * 1000,
            )

    # ── 11. template ───────────────────────────────────────────────────────────

    async def _template(self, kwargs: Dict, start_time: float) -> ToolResult:
        template_type = kwargs.get("template_type", "invoice").lower()
        data = kwargs.get("data", {})
        output_path = self._validate_path(kwargs.get("output_path", f"{template_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"))

        templates = {
            "invoice": self._template_invoice,
            "purchase_order": self._template_purchase_order,
            "timesheet": self._template_timesheet,
            "budget": self._template_budget,
        }

        builder = templates.get(template_type)
        if builder is None:
            return ToolResult.failure(
                f"Unknown template type: {template_type}. Available: {list(templates.keys())}",
                execution_time_ms=(time.time() - start_time) * 1000,
            )

        try:
            builder(output_path, data)
            elapsed = (time.time() - start_time) * 1000
            return ToolResult.success(
                f"Template '{template_type}' created: {output_path.name}",
                data={"path": str(output_path), "template_type": template_type},
                execution_time_ms=elapsed,
            )
        except Exception as e:
            return ToolResult.error_result(
                f"Template generation failed: {e}",
                error=traceback.format_exc(),
                execution_time_ms=(time.time() - start_time) * 1000,
            )

    def _template_invoice(self, path: Path, data: Dict) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoice"

        company = data.get("company", "Your Company")
        client = data.get("client", "Client Name")
        invoice_no = data.get("invoice_no", "INV-001")
        invoice_date = data.get("date", datetime.now().strftime("%d-%m-%Y"))
        due_date = data.get("due_date", (datetime.now() + timedelta(days=30)).strftime("%d-%m-%Y"))
        items = data.get("items", [])
        tax_rate = float(data.get("tax_rate", 10))
        notes = data.get("notes", "Thank you for your business!")

        dark_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        light_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

        ws.merge_cells("A1:F1")
        ws["A1"] = "INVOICE"
        ws["A1"].font = Font(bold=True, size=20, color="1F4E79")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40

        ws["A3"] = f"From: {company}"
        ws["A4"] = f"Bill To: {client}"
        ws["D3"] = f"Invoice #: {invoice_no}"
        ws["D4"] = f"Date: {invoice_date}"
        ws["D5"] = f"Due Date: {due_date}"

        headers = ["#", "Description", "Quantity", "Unit Price", "Amount", "Tax"]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=7, column=ci, value=h)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = dark_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self._style_border

        subtotal = 0.0
        for ri, item in enumerate(items, 8):
            qty = float(item.get("quantity", 1))
            price = float(item.get("unit_price", 0))
            amount = qty * price
            tax = amount * tax_rate / 100
            subtotal += amount

            vals = [ri - 7, item.get("description", f"Item {ri - 7}"), qty, price, round(amount, 2), round(tax, 2)]
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.border = self._style_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if (ri - 8) % 2 == 1:
                    cell.fill = light_fill

        total_row = 8 + len(items)
        ws.cell(row=total_row, column=4, value="Subtotal:").font = Font(bold=True)
        ws.cell(row=total_row, column=5, value=round(subtotal, 2)).font = Font(bold=True)

        tax_total = subtotal * tax_rate / 100
        ws.cell(row=total_row + 1, column=4, value=f"Tax ({tax_rate}%):").font = Font(bold=True)
        ws.cell(row=total_row + 1, column=5, value=round(tax_total, 2)).font = Font(bold=True)

        grand = subtotal + tax_total
        ws.cell(row=total_row + 2, column=4, value="Total:").font = Font(bold=True, size=14, color="1F4E79")
        ws.cell(row=total_row + 2, column=5, value=round(grand, 2)).font = Font(bold=True, size=14, color="1F4E79")

        note_row = total_row + 4
        ws.cell(row=note_row, column=1, value="Notes:").font = Font(bold=True)
        ws.cell(row=note_row + 1, column=1, value=notes)

        total_cols = 6
        for ri in range(7, total_row + 3):
            for ci in range(1, total_cols + 1):
                ws.cell(row=ri, column=ci).border = self._style_border

        self._auto_column_width(ws, total_cols, total_row + 3)
        self._freeze_header(ws, "A8")
        wb.save(str(path))

    def _template_purchase_order(self, path: Path, data: Dict) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Purchase Order"

        po_no = data.get("po_no", "PO-001")
        vendor = data.get("vendor", "Vendor Name")
        date = data.get("date", datetime.now().strftime("%d-%m-%Y"))
        items = data.get("items", [])
        notes = data.get("notes", "")

        dark_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

        ws.merge_cells("A1:F1")
        ws["A1"] = "PURCHASE ORDER"
        ws["A1"].font = Font(bold=True, size=20, color="2F5496")

        ws["A3"] = f"PO No: {po_no}"
        ws["D3"] = f"Date: {date}"
        ws["A4"] = f"Vendor: {vendor}"

        headers = ["Item #", "Description", "Quantity", "Unit Price", "Total", "Delivery Date"]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=6, column=ci, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = dark_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self._style_border

        grand_total = 0.0
        for ri, item in enumerate(items, 7):
            qty = float(item.get("quantity", 1))
            price = float(item.get("unit_price", 0))
            total = qty * price
            grand_total += total
            vals = [ri - 6, item.get("description", ""), qty, price, round(total, 2), item.get("delivery_date", "")]
            for ci, val in enumerate(vals, 1):
                ws.cell(row=ri, column=ci, value=val).border = self._style_border

        total_row = 7 + len(items)
        ws.cell(row=total_row, column=4, value="Grand Total:").font = Font(bold=True, size=12)
        ws.cell(row=total_row, column=5, value=round(grand_total, 2)).font = Font(bold=True, size=12)
        for ci in range(1, 7):
            ws.cell(row=total_row, column=ci).border = self._style_border

        if notes:
            ws.cell(row=total_row + 2, column=1, value="Notes:").font = Font(bold=True)
            ws.cell(row=total_row + 3, column=1, value=notes)

        self._auto_column_width(ws, 6, total_row + 3)
        wb.save(str(path))

    def _template_timesheet(self, path: Path, data: Dict) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Timesheet"

        employee = data.get("employee", "Employee Name")
        period = data.get("period", datetime.now().strftime("%B %Y"))
        entries = data.get("entries", [])
        hourly_rate = float(data.get("hourly_rate", 0))

        dark_fill = PatternFill(start_color="385723", end_color="385723", fill_type="solid")

        ws.merge_cells("A1:G1")
        ws["A1"] = "TIMESHEET"
        ws["A1"].font = Font(bold=True, size=20, color="385723")

        ws["A3"] = f"Employee: {employee}"
        ws["D3"] = f"Period: {period}"
        if hourly_rate:
            ws["F3"] = f"Rate: ₹{hourly_rate:,.2f}/hr"

        headers = ["Date", "Day", "Project", "Task", "Start", "End", "Hours"]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=5, column=ci, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = dark_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self._style_border

        total_hours = 0.0
        light_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

        for ri, entry in enumerate(entries, 6):
            hours = float(entry.get("hours", 0))
            total_hours += hours
            vals = [
                entry.get("date", ""),
                entry.get("day", ""),
                entry.get("project", ""),
                entry.get("task", ""),
                entry.get("start", ""),
                entry.get("end", ""),
                hours,
            ]
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.border = self._style_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if (ri - 6) % 2 == 0:
                    cell.fill = light_fill

        total_row = 6 + len(entries)
        ws.cell(row=total_row, column=6, value="Total Hours:").font = Font(bold=True)
        ws.cell(row=total_row, column=7, value=total_hours).font = Font(bold=True)
        for ci in range(1, 8):
            ws.cell(row=total_row, column=ci).border = self._style_border

        if hourly_rate:
            ws.cell(row=total_row + 1, column=6, value="Total Pay:")
            ws.cell(row=total_row + 1, column=7, value=round(total_hours * hourly_rate, 2))
            ws.cell(row=total_row + 1, column=7).number_format = '₹#,##0.00'

        ws.cell(row=total_row + 2, column=7, value=f"=SUM(G6:G{total_row - 1})")
        ws.cell(row=total_row + 2, column=7).value = None

        self._auto_column_width(ws, 7, total_row + 2)
        wb.save(str(path))

    def _template_budget(self, path: Path, data: Dict) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Budget"

        title = data.get("title", "Budget Report")
        period = data.get("period", f"FY {datetime.now().year}")
        categories = data.get("categories", [])

        dark_fill = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        ws.merge_cells("A1:D1")
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=18, color="BF8F00")

        ws["A3"] = f"Period: {period}"

        headers = ["Category", "Budgeted", "Actual", "Variance"]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=5, column=ci, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = dark_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self._style_border

        total_budgeted = 0.0
        total_actual = 0.0

        for ri, cat in enumerate(categories, 6):
            budgeted = float(cat.get("budgeted", 0))
            actual = float(cat.get("actual", 0))
            variance = actual - budgeted
            total_budgeted += budgeted
            total_actual += actual

            vals = [cat.get("name", ""), round(budgeted, 2), round(actual, 2), round(variance, 2)]
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.border = self._style_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if variance < 0:
                    cell.fill = red_fill
                elif variance > 0:
                    cell.fill = green_fill

        total_row = 6 + len(categories)
        ws.cell(row=total_row, column=1, value="Total").font = Font(bold=True)
        ws.cell(row=total_row, column=2, value=round(total_budgeted, 2)).font = Font(bold=True)
        ws.cell(row=total_row, column=3, value=round(total_actual, 2)).font = Font(bold=True)
        ws.cell(row=total_row, column=4, value=round(total_actual - total_budgeted, 2)).font = Font(bold=True)
        for ci in range(1, 5):
            ws.cell(row=total_row, column=ci).border = self._style_border

        summary_row = total_row + 2
        ws.cell(row=summary_row, column=1, value="Summary:").font = Font(bold=True)
        ws.cell(row=summary_row + 1, column=1, value="Total Budgeted")
        ws.cell(row=summary_row + 1, column=2, value=round(total_budgeted, 2))
        ws.cell(row=summary_row + 2, column=1, value="Total Actual")
        ws.cell(row=summary_row + 2, column=2, value=round(total_actual, 2))
        ws.cell(row=summary_row + 3, column=1, value="Net Variance")
        ws.cell(row=summary_row + 3, column=2, value=round(total_actual - total_budgeted, 2))

        self._auto_column_width(ws, 4, summary_row + 3)
        ws.cell(row=summary_row + 3, column=1).font = Font(bold=True, size=12)
        wb.save(str(path))

    # ── 12. merge_workbooks ────────────────────────────────────────────────────

    async def _merge_workbooks(self, kwargs: Dict, start_time: float) -> ToolResult:
        files = kwargs.get("files", [])
        output_path = self._validate_path(kwargs.get("output_path", "merged_workbook.xlsx"))
        suffix_strategy = kwargs.get("suffix_strategy", "number")
        include_summary = kwargs.get("include_summary", True)

        if not files:
            file_pattern = kwargs.get("file_pattern", "")
            if file_pattern:
                import glob as glob_module
                files = glob_module.glob(file_pattern)
            else:
                return ToolResult.failure(
                    "merge_workbooks requires 'files' (list of file paths) or 'file_pattern' (glob)",
                    execution_time_ms=(time.time() - start_time) * 1000,
                )

        if not files:
            return ToolResult.failure(
                "No files matched the given pattern or list",
                execution_time_ms=(time.time() - start_time) * 1000,
            )

        try:
            out_wb = Workbook()
            out_wb.remove(out_wb.active)
            merged_sheets: Dict[str, int] = {}
            summary_info = []

            for file_idx, fp in enumerate(files):
                fpath = Path(fp)
                if not fpath.exists():
                    continue

                try:
                    in_wb = load_workbook(fpath, data_only=True)
                except Exception:
                    continue

                for sn in in_wb.sheetnames:
                    ws = in_wb[sn]

                    target_name = sn
                    if target_name in merged_sheets:
                        if suffix_strategy == "number":
                            target_name = f"{sn}_{file_idx + 1}"
                        elif suffix_strategy == "file":
                            target_name = f"{sn}_{fpath.stem}"
                        else:
                            target_name = f"{sn}_{merged_sheets.get(sn, 0) + 1}"

                    new_ws = out_wb.create_sheet(title=target_name)
                    merged_sheets[target_name] = merged_sheets.get(target_name, 0) + 1

                    for row in ws.iter_rows():
                        for cell in row:
                            new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                            if cell.has_style:
                                new_cell.font = copy(cell.font)
                                new_cell.fill = copy(cell.fill)
                                new_cell.border = copy(cell.border)
                                new_cell.alignment = copy(cell.alignment)
                                new_cell.number_format = cell.number_format

                    summary_info.append({
                        "file": fpath.name,
                        "sheet": sn,
                        "copied_as": target_name,
                        "rows": ws.max_row or 0,
                        "columns": ws.max_column or 0,
                    })

                in_wb.close()

            if include_summary and summary_info:
                summary_ws = out_wb.create_sheet(title="Merge Summary", index=0)
                sum_headers = ["File", "Original Sheet", "Copied As", "Rows", "Columns"]
                for ci, h in enumerate(sum_headers, 1):
                    summary_ws.cell(row=1, column=ci, value=h)
                self._apply_header_style(summary_ws, len(sum_headers))

                for ri, info in enumerate(summary_info, 2):
                    for ci, key in enumerate(["file", "sheet", "copied_as", "rows", "columns"], 1):
                        summary_ws.cell(row=ri, column=ci, value=info[key]).border = self._style_border

                self._auto_column_width(summary_ws, len(sum_headers), len(summary_info) + 1)

            out_wb.save(str(output_path))
            out_wb.close()

            total_sheets = len(out_wb.sheetnames) if False else len(summary_info)
            elapsed = (time.time() - start_time) * 1000
            return ToolResult.success(
                f"Merged {len(files)} file(s) into {output_path.name}: {len(set(i['file'] for i in summary_info))} files, {len(summary_info)} sheets",
                data={
                    "output_path": str(output_path),
                    "files_merged": len(files),
                    "total_sheets": len(summary_info),
                    "unique_files": len(set(i["file"] for i in summary_info)),
                    "sheets_summary": summary_info,
                },
                execution_time_ms=elapsed,
            )
        except Exception as e:
            return ToolResult.error_result(
                f"Merge workbooks failed: {e}",
                error=traceback.format_exc(),
                execution_time_ms=(time.time() - start_time) * 1000,
            )
